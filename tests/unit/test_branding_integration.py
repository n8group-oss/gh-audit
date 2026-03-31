"""Tests for N8 Group branding integration across HTML reports, summary HTML, and Excel.

Verifies that branding constants from gh_audit.branding are embedded in all
customer-facing outputs: discovery HTML report, cross-org summary HTML, and
Excel workbook Summary sheet.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from gh_audit.models.inventory import Inventory, InventoryMetadata, InventorySummary
from gh_audit.models.multi_org import MultiOrgSummary, OrgScanResult
from gh_audit.models.user import OrgMemberSummary
from gh_audit.services.excel_export import ExcelExportService
from gh_audit.services.reporting import ReportService
from gh_audit.services.summary_report import generate_summary_html


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _mock_inventory() -> Inventory:
    return Inventory(
        metadata=InventoryMetadata(
            schema_version="1.0.0",
            generated_at=datetime(2026, 3, 28, tzinfo=timezone.utc),
            tool_version="0.1.0",
            organization="testorg",
            auth_method="pat",
            scan_profile="standard",
        ),
        summary=InventorySummary(total_repos=1),
        repositories=[],
        users=OrgMemberSummary(total=1),
    )


def _mock_summary() -> MultiOrgSummary:
    return MultiOrgSummary(
        tool_version="0.1.0",
        config_file="test.yml",
        organizations=[OrgScanResult(name="testorg", status="success", total_repos=10)],
    )


# ---------------------------------------------------------------------------
# HTML report branding
# ---------------------------------------------------------------------------


class TestHtmlReportBranding:
    """Discovery HTML report contains N8 Group branding elements."""

    def _render(self, tmp_path: Path) -> str:
        inv = _mock_inventory()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        return output.read_text(encoding="utf-8")

    def test_contains_about_n8_group(self, tmp_path: Path) -> None:
        content = self._render(tmp_path)
        assert "About N8 Group" in content

    def test_contains_website(self, tmp_path: Path) -> None:
        content = self._render(tmp_path)
        assert "n8-group.com" in content

    def test_contains_sales_email(self, tmp_path: Path) -> None:
        content = self._render(tmp_path)
        assert "sales@n8-group.com" in content

    def test_contains_service_names(self, tmp_path: Path) -> None:
        content = self._render(tmp_path)
        # At least some services from the branding module should appear
        assert "GitHub Enterprise Governance" in content
        assert "DevOps Strategy Consulting" in content

    def test_footer_mentions_free_tool(self, tmp_path: Path) -> None:
        content = self._render(tmp_path)
        assert "free tool" in content.lower()

    def test_footer_mentions_gh_audit(self, tmp_path: Path) -> None:
        content = self._render(tmp_path)
        assert "gh-audit" in content

    def test_footer_mentions_n8_group(self, tmp_path: Path) -> None:
        content = self._render(tmp_path)
        assert "N8 Group" in content


# ---------------------------------------------------------------------------
# Summary HTML branding
# ---------------------------------------------------------------------------


class TestSummaryHtmlBranding:
    """Cross-org summary HTML report contains N8 Group branding elements."""

    def _render(self, tmp_path: Path) -> str:
        summary = _mock_summary()
        output = tmp_path / "summary.html"
        generate_summary_html(summary, output)
        return output.read_text(encoding="utf-8")

    def test_contains_about_n8_group(self, tmp_path: Path) -> None:
        content = self._render(tmp_path)
        assert "About N8 Group" in content

    def test_contains_website(self, tmp_path: Path) -> None:
        content = self._render(tmp_path)
        assert "n8-group.com" in content

    def test_footer_mentions_gh_audit(self, tmp_path: Path) -> None:
        content = self._render(tmp_path)
        assert "gh-audit" in content


# ---------------------------------------------------------------------------
# Excel branding
# ---------------------------------------------------------------------------


class TestExcelBranding:
    """Excel Summary sheet contains N8 Group branding rows."""

    def _generate(self, tmp_path: Path) -> Path:
        inv = _mock_inventory()
        output = tmp_path / "workbook.xlsx"
        ExcelExportService.generate(inv, output)
        return output

    def test_summary_sheet_contains_n8_group(self, tmp_path: Path) -> None:
        wb_path = self._generate(tmp_path)
        wb = load_workbook(str(wb_path))
        ws = wb["Summary"]
        # Collect all cell values as strings
        values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    values.append(str(cell))
        text = " ".join(values)
        assert "N8 Group" in text

    def test_summary_sheet_contains_website(self, tmp_path: Path) -> None:
        wb_path = self._generate(tmp_path)
        wb = load_workbook(str(wb_path))
        ws = wb["Summary"]
        values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    values.append(str(cell))
        text = " ".join(values)
        assert "n8-group.com" in text

    def test_summary_sheet_contains_gh_audit(self, tmp_path: Path) -> None:
        wb_path = self._generate(tmp_path)
        wb = load_workbook(str(wb_path))
        ws = wb["Summary"]
        values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    values.append(str(cell))
        text = " ".join(values)
        assert "gh-audit" in text
