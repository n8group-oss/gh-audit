"""ExcelExportService — generates a polished multi-sheet Excel workbook from an Inventory.

Sheet contract (10 core sheets, in order):
1.  Summary       — branded header, org metadata, high-level stats
2.  Repositories  — one row per repo with core fields
3.  Actions       — one row per workflow
4.  Security      — per-repo security enablement and alert counts
5.  Issues        — per-repo issue counts and top labels
6.  Packages      — name, type, visibility
7.  Projects      — title, item_count, closed (Yes/No)
8.  Users         — role, count
9.  Large Files   — repo name, file path, size_mb
10. Warnings      — repo name (or "Scan"), warning text

Partial-scan semantics:
- bool | None fields -> "Yes" / "No" / "n/a"
- int  | None counts -> number / "n/a"

Formatting:
- Professional styling matching ado2gh: dark-blue headers, Excel tables,
  alternating rows, thin borders, auto-fit columns, tab colours, number
  formatting, and a branded Summary sheet with coloured sections.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from gh_audit import branding
from gh_audit.models.inventory import Inventory

# ---------------------------------------------------------------------------
# Style constants  (matching ado2gh colour palette)
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_METADATA_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_METADATA_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")

_TOTALS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_TOTALS_FONT = Font(name="Calibri", size=11, bold=True)

_DATA_FONT = Font(name="Calibri", size=10)
_BOLD_DATA_FONT = Font(name="Calibri", size=10, bold=True)
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

_ALT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 50

_NUMBER_FMT = "#,##0"
_DECIMAL_FMT = "#,##0.00"

_TAB_COLORS: dict[str, str] = {
    "Summary": "1F4E79",
    "Repositories": "2E75B6",
    "Actions": "2E75B6",
    "Security": "C00000",
    "Issues": "BF8F00",
    "Packages": "548235",
    "Projects": "548235",
    "Users": "7F7F7F",
    "Large Files": "7F7F7F",
    "Warnings": "C00000",
    "Teams": "2E75B6",
    "Org Policies": "2E75B6",
    "Org Rulesets": "2E75B6",
    "Runners": "BF8F00",
    "Environments": "BF8F00",
    "Installed Apps": "BF8F00",
    "Dependabot Alerts": "C00000",
    "Code Scanning Alerts": "C00000",
    "Secret Scanning Alerts": "C00000",
    "Copilot": "548235",
    "Traffic": "548235",
    "Community Health": "548235",
    "Actions Runs": "BF8F00",
    "Enterprise": "1F4E79",
    "Enterprise Teams": "1F4E79",
}


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _bool_na(value: bool | None) -> str:
    """Convert bool | None to 'Yes' / 'No' / 'n/a'."""
    if value is None:
        return "n/a"
    return "Yes" if value else "No"


def _int_na(value: int | None) -> int | str:
    """Return the integer value or 'n/a' when None."""
    if value is None:
        return "n/a"
    return value


def _style_header_row(ws: Worksheet, headers: list[str]) -> None:
    """Write the header row with dark-blue fill, white bold font, and freeze pane."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    # Auto-width from header lengths (refined later by _auto_fit_columns)
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(header) + 2, _MIN_COL_WIDTH)

    # Freeze the header row
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_summary(ws: Worksheet, inventory: Inventory) -> None:
    """Summary sheet: branded header, metadata section, totals section, branding."""
    meta = inventory.metadata
    summary = inventory.summary

    # --- Branding block (rows 1-3) ---
    ws.append(["gh-audit Discovery Report"])
    ws.append([branding.TAGLINE])
    ws.append(["github.com/n8group-oss/gh-audit"])

    # Blank separator (row 4)
    ws.append([None])

    # Header row (row 5)
    ws.append(["Section", "Key", "Value"])

    # --- Metadata section ---
    metadata_rows = [
        ("Metadata", "Organization", meta.organization),
        ("Metadata", "Tool Version", meta.tool_version),
        ("Metadata", "Schema Version", meta.schema_version),
        ("Metadata", "Generated At", str(meta.generated_at)),
        ("Metadata", "Auth Method", meta.auth_method),
        ("Metadata", "Scan Profile", meta.scan_profile),
        ("Metadata", "API URL", meta.api_url),
    ]
    for row in metadata_rows:
        ws.append(list(row))

    # --- Totals section ---
    totals_rows = [
        ("Totals", "Total Repositories", summary.total_repos),
        ("Totals", "Public Repositories", summary.public_repos),
        ("Totals", "Private Repositories", summary.private_repos),
        ("Totals", "Internal Repositories", summary.internal_repos),
        ("Totals", "Archived Repositories", summary.archived_repos),
        ("Totals", "Forked Repositories", summary.forked_repos),
        ("Totals", "Template Repositories", summary.template_repos),
        ("Totals", "Total Size (bytes)", summary.total_size_bytes),
        ("Totals", "Total Branches", summary.total_branches),
        ("Totals", "Total PRs", summary.total_prs),
        ("Totals", "Total Issues", summary.total_issues),
        ("Totals", "Repos with Large Files", summary.repos_with_large_files),
        ("Totals", "Repos with LFS", summary.repos_with_lfs),
        ("Totals", "Repos with Workflows", summary.repos_with_workflows),
        ("Totals", "Total Workflows", summary.total_workflow_count),
        ("Totals", "Repos with Self-Hosted Runners", summary.repos_with_self_hosted_runners),
        ("Totals", "Repos with Dependabot", summary.repos_with_dependabot),
        ("Totals", "Repos with Code Scanning", summary.repos_with_code_scanning),
        ("Totals", "Repos with Secret Scanning", summary.repos_with_secret_scanning),
        ("Totals", "Total Packages", summary.total_packages),
        ("Totals", "Total Projects", summary.total_projects),
    ]
    for row in totals_rows:
        ws.append(list(row))

    # Blank separator then N8 Group branding
    ws.append([None])
    branding_rows = [
        ("About", branding.COMPANY_NAME, f"gh-audit is a free tool by {branding.COMPANY_NAME}"),
        ("About", "Website", branding.WEBSITE),
        ("About", "Contact", branding.SALES_EMAIL),
        ("About", "Phone", branding.PHONE),
        ("About", "Services", ", ".join(branding.SERVICES)),
    ]
    for row in branding_rows:
        ws.append(list(row))


def _build_repositories(ws: Worksheet, inventory: Inventory) -> None:
    headers = [
        "Repository",
        "Visibility",
        "Language",
        "Size (bytes)",
        "Branches",
        "Open PRs",
        "Merged PRs",
        "Open Issues",
        "Closed Issues",
        "Workflows",
        "Archived",
        "Fork",
        "Template",
        "LFS",
        "Large Files Count",
    ]
    _style_header_row(ws, headers)

    for repo in inventory.repositories:
        ws.append(
            [
                repo.name,
                repo.visibility,
                repo.language or "",
                repo.size_bytes,
                repo.branch_count,
                repo.pr_count_open,
                repo.pr_count_merged,
                repo.issue_count_open,
                repo.issue_count_closed,
                repo.actions.workflow_count,
                _bool_na(repo.archived),
                _bool_na(repo.fork),
                _bool_na(repo.is_template),
                _bool_na(repo.lfs_info.has_lfs),
                len(repo.large_file_scan.files),
            ]
        )


def _build_actions(ws: Worksheet, inventory: Inventory) -> None:
    headers = ["Repository", "Workflow Name", "Path", "State"]
    _style_header_row(ws, headers)

    for repo in inventory.repositories:
        for wf in repo.actions.workflows:
            ws.append([repo.name, wf.name, wf.path, wf.state])


def _build_security(ws: Worksheet, inventory: Inventory) -> None:
    headers = [
        "Repository",
        "Dependabot Enabled",
        "Dependabot Alerts",
        "Code Scanning Enabled",
        "Code Scanning Alerts",
        "Secret Scanning Enabled",
        "Secret Scanning Alerts",
    ]
    _style_header_row(ws, headers)

    for repo in inventory.repositories:
        sec = repo.security
        ws.append(
            [
                repo.name,
                _bool_na(sec.dependabot_enabled),
                _int_na(sec.dependabot_alerts_open),
                _bool_na(sec.code_scanning_enabled),
                _int_na(sec.code_scanning_alerts_open),
                _bool_na(sec.secret_scanning_enabled),
                _int_na(sec.secret_scanning_alerts_open),
            ]
        )


def _build_issues(ws: Worksheet, inventory: Inventory) -> None:
    # Gather all label keys across repos for extra columns
    all_labels: list[str] = []
    seen: set[str] = set()
    for repo in inventory.repositories:
        for label in repo.issue_label_distribution:
            if label not in seen:
                all_labels.append(label)
                seen.add(label)

    headers = ["Repository", "Open Issues", "Closed Issues"] + all_labels
    _style_header_row(ws, headers)

    for repo in inventory.repositories:
        label_counts = [repo.issue_label_distribution.get(lbl, 0) for lbl in all_labels]
        ws.append([repo.name, repo.issue_count_open, repo.issue_count_closed] + label_counts)


def _build_packages(ws: Worksheet, inventory: Inventory) -> None:
    headers = ["Name", "Type", "Visibility"]
    _style_header_row(ws, headers)

    for pkg in inventory.packages:
        ws.append([pkg.name, pkg.package_type, pkg.visibility])


def _build_projects(ws: Worksheet, inventory: Inventory) -> None:
    headers = ["Title", "Item Count", "Closed"]
    _style_header_row(ws, headers)

    for proj in inventory.projects:
        ws.append([proj.title, proj.item_count, _bool_na(proj.closed)])


def _build_users(ws: Worksheet, inventory: Inventory) -> None:
    headers = ["Role", "Count"]
    _style_header_row(ws, headers)

    users = inventory.users
    ws.append(["Admin", users.admins])
    ws.append(["Member", users.members])
    ws.append(["Outside Collaborator", users.outside_collaborators])
    ws.append(["Total", users.total])


def _build_large_files(ws: Worksheet, inventory: Inventory) -> None:
    headers = ["Repository", "File Path", "Size (MB)"]
    _style_header_row(ws, headers)

    for repo in inventory.repositories:
        for lf in repo.large_file_scan.files:
            size_mb = round(lf.size_bytes / (1024 * 1024), 2)
            ws.append([repo.name, lf.path, size_mb])


def _build_warnings(ws: Worksheet, inventory: Inventory) -> None:
    headers = ["Source", "Warning"]
    _style_header_row(ws, headers)

    # Scan-level warnings first
    for warning in inventory.metadata.scan_warnings:
        ws.append(["Scan", warning])

    # Per-repo warnings
    for repo in inventory.repositories:
        for warning in repo.warnings:
            ws.append([repo.name, warning])


# ---------------------------------------------------------------------------
# Formatting helpers  (applied after all data is written)
# ---------------------------------------------------------------------------


def _add_table(sheet: Worksheet, name: str) -> None:
    """Add an Excel table over all data in the sheet (header + data rows)."""
    max_row = sheet.max_row
    max_col = sheet.max_column
    if max_row < 1 or max_col < 1:
        return
    # Tables require at least 2 rows (header + 1 data)
    if max_row == 1:
        sheet.append([None] * max_col)
        max_row = 2
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _auto_fit_columns(
    sheet: Worksheet, min_width: int = _MIN_COL_WIDTH, max_width: int = _MAX_COL_WIDTH
) -> None:
    """Set each column width to fit the longest cell value."""
    for col_cells in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = max(len(str(line)) for line in str(cell.value).split("\n"))
                max_len = max(max_len, cell_len)
        adjusted = min(max(max_len + 2, min_width), max_width)
        sheet.column_dimensions[col_letter].width = adjusted


def _style_data_cells(sheet: Worksheet, start_row: int = 2) -> None:
    """Apply font, border, and wrap alignment to all data cells."""
    for row in sheet.iter_rows(
        min_row=start_row, max_row=sheet.max_row, max_col=sheet.max_column
    ):
        for cell in row:
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _WRAP_ALIGNMENT


def _bold_first_column(sheet: Worksheet) -> None:
    """Bold the first column (entity names) in data rows."""
    for row in sheet.iter_rows(min_row=2):
        if row[0].value is not None:
            row[0].font = _BOLD_DATA_FONT


def _apply_alternating_rows(sheet: Worksheet, start_row: int = 3) -> None:
    """Apply subtle alternating row shading to data rows."""
    for row_idx in range(start_row, sheet.max_row + 1, 2):
        for cell in sheet[row_idx]:
            if cell.fill == PatternFill():  # only if not already coloured
                cell.fill = _ALT_ROW_FILL


def _apply_number_formatting(sheet: Worksheet) -> None:
    """Apply thousands separator to integer cells and 2-decimal to floats."""
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = _NUMBER_FMT
            elif isinstance(cell.value, float):
                cell.number_format = _DECIMAL_FMT


def _re_style_header_row(sheet: Worksheet) -> None:
    """Re-apply header styling after table creation (tables can override styles)."""
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _style_summary_sheet(ws: Worksheet) -> None:
    """Apply visual sections to the Summary sheet."""
    # --- Branding rows (1-3) ---
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:C2")
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="808080")

    ws.merge_cells("A3:C3")
    ws["A3"].hyperlink = "https://github.com/n8group-oss/gh-audit"
    ws["A3"].font = Font(name="Calibri", size=10, color="2E75B6", underline="single")
    ws["A3"].value = "github.com/n8group-oss/gh-audit"

    # --- Row 5: header row ---
    for cell in ws[5]:
        if cell.value is not None:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER

    # --- Data rows: section-based colouring ---
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=ws.max_column):
        section = row[0].value
        for cell in row:
            cell.border = _THIN_BORDER

        if section == "Metadata":
            for cell in row:
                cell.fill = _METADATA_FILL
                cell.font = _METADATA_FONT
        elif section == "Totals":
            for cell in row:
                cell.fill = _TOTALS_FILL
                cell.font = _TOTALS_FONT
            # Right-align the value column (C) for totals
            if len(row) >= 3 and row[2].value is not None:
                row[2].alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(row[2].value, (int, float)):
                    row[2].number_format = _NUMBER_FMT
        elif section == "About":
            for cell in row:
                cell.font = Font(name="Calibri", size=10, color="808080")

    # Freeze below branding + header
    ws.freeze_panes = "A6"

    # Auto-fit columns
    _auto_fit_columns(ws)


def _apply_formatting(wb: Workbook) -> None:
    """Apply comprehensive formatting: tables, styles, colours to the entire workbook."""
    # --- Summary sheet ---
    if "Summary" in wb.sheetnames:
        _style_summary_sheet(wb["Summary"])

    # --- Data sheets ---
    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue
        sheet = wb[sheet_name]

        # Add Excel table (gives row stripes and filter dropdowns)
        safe_name = sheet_name.replace(" ", "")
        _add_table(sheet, safe_name)

        # Re-apply header style after table creation
        _re_style_header_row(sheet)

        # Style data cells (font, border, wrap)
        _style_data_cells(sheet)

        # Alternating row shading
        _apply_alternating_rows(sheet)

        # Bold first column (entity names)
        _bold_first_column(sheet)

        # Number formatting
        _apply_number_formatting(sheet)

        # Auto-fit columns
        _auto_fit_columns(sheet)

    # --- Tab colours ---
    for sheet_name, color in _TAB_COLORS.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


class ExcelExportService:
    """Generate a multi-sheet Excel workbook from a completed Inventory."""

    @staticmethod
    def generate(inventory: Inventory, output_path: Path) -> None:
        """Write the Excel workbook to *output_path*.

        Parameters
        ----------
        inventory:
            Complete scan inventory.
        output_path:
            Destination ``.xlsx`` file path.  Parent directories are created
            automatically.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Remove the default empty sheet created by openpyxl
        wb.remove(wb.active)  # type: ignore[arg-type]

        # Build sheets in required order
        _build_summary(wb.create_sheet("Summary"), inventory)
        _build_repositories(wb.create_sheet("Repositories"), inventory)
        _build_actions(wb.create_sheet("Actions"), inventory)
        _build_security(wb.create_sheet("Security"), inventory)
        _build_issues(wb.create_sheet("Issues"), inventory)
        _build_packages(wb.create_sheet("Packages"), inventory)
        _build_projects(wb.create_sheet("Projects"), inventory)
        _build_users(wb.create_sheet("Users"), inventory)
        _build_large_files(wb.create_sheet("Large Files"), inventory)
        _build_warnings(wb.create_sheet("Warnings"), inventory)

        # Governance sheets (only when governance data present)
        if inventory.governance is not None:
            gov = inventory.governance

            # Teams sheet
            ws = wb.create_sheet("Teams")
            teams_headers = ["Team", "Permission", "Privacy", "Members", "Repos", "Parent"]
            _style_header_row(ws, teams_headers)
            for idx, team in enumerate(gov.teams, 2):
                ws.cell(row=idx, column=1, value=team.name)
                ws.cell(row=idx, column=2, value=team.permission)
                ws.cell(row=idx, column=3, value=team.privacy)
                ws.cell(row=idx, column=4, value=team.member_count)
                ws.cell(row=idx, column=5, value=team.repo_count)
                ws.cell(row=idx, column=6, value=team.parent_team or "")

            # Org Policies sheet
            ws = wb.create_sheet("Org Policies")
            pol_headers = ["Policy", "Value"]
            _style_header_row(ws, pol_headers)
            pol = gov.org_policies
            pol_rows = [
                ("Default repo permission", pol.default_repository_permission),
                ("2FA required", _bool_na(pol.two_factor_requirement_enabled)),
                ("Web commit sign-off", _bool_na(pol.web_commit_signoff_required)),
                ("Members can create repos", _bool_na(pol.members_can_create_repositories)),
                (
                    "Members can fork private",
                    _bool_na(pol.members_can_fork_private_repositories),
                ),
                ("Members can delete repos", _bool_na(pol.members_can_delete_repositories)),
                (
                    "Members can change visibility",
                    _bool_na(pol.members_can_change_repo_visibility),
                ),
            ]
            for idx, (policy, value) in enumerate(pol_rows, 2):
                ws.cell(row=idx, column=1, value=policy)
                ws.cell(row=idx, column=2, value=value if value is not None else "n/a")

            # Org Rulesets sheet
            ws = wb.create_sheet("Org Rulesets")
            rs_headers = ["Name", "Enforcement", "Target", "Source", "Rules Count"]
            _style_header_row(ws, rs_headers)
            for idx, rs in enumerate(gov.org_rulesets, 2):
                ws.cell(row=idx, column=1, value=rs.name)
                ws.cell(row=idx, column=2, value=rs.enforcement)
                ws.cell(row=idx, column=3, value=rs.target)
                ws.cell(row=idx, column=4, value=rs.source_type)
                ws.cell(row=idx, column=5, value=len(rs.rules))

        # Operations sheets (only when operations data present)
        if inventory.operations is not None:
            ops = inventory.operations

            # Runners sheet
            ws = wb.create_sheet("Runners")
            runners_headers = ["Name", "OS", "Status", "Labels", "Group"]
            _style_header_row(ws, runners_headers)
            for idx, runner in enumerate(ops.runners, 2):
                ws.cell(row=idx, column=1, value=runner.name)
                ws.cell(row=idx, column=2, value=runner.os)
                ws.cell(row=idx, column=3, value=runner.status)
                ws.cell(row=idx, column=4, value=", ".join(runner.labels))
                ws.cell(row=idx, column=5, value=runner.runner_group_name or "")

            # Environments sheet
            ws = wb.create_sheet("Environments")
            env_headers = [
                "Repository",
                "Environment",
                "Wait Timer",
                "Required Reviewers",
                "Branch Policy",
                "Can Admins Bypass",
            ]
            _style_header_row(ws, env_headers)
            env_row = 2
            for repo in inventory.repositories:
                if repo.environments:
                    for env in repo.environments:
                        ws.cell(row=env_row, column=1, value=repo.name)
                        ws.cell(row=env_row, column=2, value=env.name)
                        prot = env.protection_rules
                        ws.cell(row=env_row, column=3, value=prot.wait_timer if prot else 0)
                        ws.cell(
                            row=env_row,
                            column=4,
                            value=prot.required_reviewers if prot else 0,
                        )
                        ws.cell(
                            row=env_row,
                            column=5,
                            value=prot.branch_policy
                            if prot and prot.branch_policy
                            else "none",
                        )
                        ws.cell(
                            row=env_row,
                            column=6,
                            value=_bool_na(env.can_admins_bypass),
                        )
                        env_row += 1

            # Installed Apps sheet
            ws = wb.create_sheet("Installed Apps")
            apps_headers = ["Name", "Slug", "Repo Selection", "Events Count"]
            _style_header_row(ws, apps_headers)
            for idx, app in enumerate(ops.installed_apps, 2):
                ws.cell(row=idx, column=1, value=app.app_name)
                ws.cell(row=idx, column=2, value=app.app_slug)
                ws.cell(row=idx, column=3, value=app.repository_selection)
                ws.cell(row=idx, column=4, value=len(app.events))

        # Security detail sheets (only when any repo has security_detail)
        has_security_detail = any(r.security_detail is not None for r in inventory.repositories)
        if has_security_detail:
            # Dependabot Alerts sheet
            ws = wb.create_sheet("Dependabot Alerts")
            dep_headers = [
                "Repository",
                "Severity",
                "Package",
                "Manifest",
                "State",
                "GHSA ID",
                "CVE",
                "Fixed Version",
            ]
            _style_header_row(ws, dep_headers)
            dep_row = 2
            for repo in inventory.repositories:
                if repo.security_detail:
                    for alert in repo.security_detail.dependabot_alerts:
                        ws.cell(row=dep_row, column=1, value=repo.name)
                        ws.cell(row=dep_row, column=2, value=alert.severity)
                        ws.cell(row=dep_row, column=3, value=alert.package_name)
                        ws.cell(row=dep_row, column=4, value=alert.manifest_path)
                        ws.cell(row=dep_row, column=5, value=alert.state)
                        ws.cell(row=dep_row, column=6, value=alert.ghsa_id or "")
                        ws.cell(row=dep_row, column=7, value=alert.cve_id or "")
                        ws.cell(row=dep_row, column=8, value=alert.fixed_version or "")
                        dep_row += 1

            # Code Scanning Alerts sheet
            ws = wb.create_sheet("Code Scanning Alerts")
            cs_headers = [
                "Repository",
                "Rule",
                "Severity",
                "Security Severity",
                "Tool",
                "State",
            ]
            _style_header_row(ws, cs_headers)
            cs_row = 2
            for repo in inventory.repositories:
                if repo.security_detail:
                    for alert in repo.security_detail.code_scanning_alerts:
                        ws.cell(row=cs_row, column=1, value=repo.name)
                        ws.cell(row=cs_row, column=2, value=alert.rule_id)
                        ws.cell(row=cs_row, column=3, value=alert.severity or "")
                        ws.cell(row=cs_row, column=4, value=alert.security_severity or "")
                        ws.cell(row=cs_row, column=5, value=alert.tool_name)
                        ws.cell(row=cs_row, column=6, value=alert.state)
                        cs_row += 1

            # Secret Scanning Alerts sheet
            ws = wb.create_sheet("Secret Scanning Alerts")
            ss_headers = [
                "Repository",
                "Secret Type",
                "State",
                "Resolution",
                "Push Protection Bypassed",
            ]
            _style_header_row(ws, ss_headers)
            ss_row = 2
            for repo in inventory.repositories:
                if repo.security_detail:
                    for alert in repo.security_detail.secret_scanning_alerts:
                        ws.cell(row=ss_row, column=1, value=repo.name)
                        ws.cell(
                            row=ss_row,
                            column=2,
                            value=alert.secret_type_display_name or alert.secret_type,
                        )
                        ws.cell(row=ss_row, column=3, value=alert.state)
                        ws.cell(row=ss_row, column=4, value=alert.resolution or "")
                        ws.cell(
                            row=ss_row,
                            column=5,
                            value=_bool_na(alert.push_protection_bypassed),
                        )
                        ss_row += 1

        # Adoption sheets (only when adoption data present)
        if inventory.adoption is not None:
            adoption = inventory.adoption

            # Copilot sheet (only if copilot data is available)
            if adoption.copilot is not None:
                ws = wb.create_sheet("Copilot")
                cop_headers = ["Metric", "Value"]
                _style_header_row(ws, cop_headers)
                cop_rows = [
                    ("Total Seats", adoption.copilot.total_seats),
                    ("Active Seats", _int_na(adoption.copilot.active_seats)),
                    ("Suggestions (28d)", _int_na(adoption.copilot.suggestions_count)),
                    ("Acceptances (28d)", _int_na(adoption.copilot.acceptances_count)),
                    (
                        "Top Languages",
                        ", ".join(adoption.copilot.top_languages)
                        if adoption.copilot.top_languages
                        else "n/a",
                    ),
                ]
                for idx, (metric, value) in enumerate(cop_rows, 2):
                    ws.cell(row=idx, column=1, value=metric)
                    ws.cell(row=idx, column=2, value=value)

            # Traffic sheet
            traffic_repos = [r for r in inventory.repositories if r.traffic is not None]
            if traffic_repos:
                ws = wb.create_sheet("Traffic")
                traffic_headers = [
                    "Repository",
                    "Views",
                    "Unique Visitors",
                    "Clones",
                    "Unique Cloners",
                ]
                _style_header_row(ws, traffic_headers)
                for idx, repo in enumerate(traffic_repos, 2):
                    t = repo.traffic
                    ws.cell(row=idx, column=1, value=repo.name)
                    ws.cell(row=idx, column=2, value=_int_na(t.views_14d))
                    ws.cell(row=idx, column=3, value=_int_na(t.unique_visitors_14d))
                    ws.cell(row=idx, column=4, value=_int_na(t.clones_14d))
                    ws.cell(row=idx, column=5, value=_int_na(t.unique_cloners_14d))

            # Community Health sheet
            health_repos = [r for r in inventory.repositories if r.community_profile is not None]
            if health_repos:
                ws = wb.create_sheet("Community Health")
                health_headers = [
                    "Repository",
                    "Health %",
                    "README",
                    "License",
                    "Contributing",
                    "Code of Conduct",
                    "Issue Template",
                    "PR Template",
                ]
                _style_header_row(ws, health_headers)
                for idx, repo in enumerate(health_repos, 2):
                    cp = repo.community_profile
                    ws.cell(row=idx, column=1, value=repo.name)
                    ws.cell(row=idx, column=2, value=cp.health_percentage)
                    ws.cell(row=idx, column=3, value=_bool_na(cp.has_readme))
                    ws.cell(row=idx, column=4, value=_bool_na(cp.has_license))
                    ws.cell(row=idx, column=5, value=_bool_na(cp.has_contributing))
                    ws.cell(row=idx, column=6, value=_bool_na(cp.has_code_of_conduct))
                    ws.cell(row=idx, column=7, value=_bool_na(cp.has_issue_template))
                    ws.cell(row=idx, column=8, value=_bool_na(cp.has_pull_request_template))

            # Actions Runs sheet
            runs_repos = [r for r in inventory.repositories if r.actions_run_summary is not None]
            if runs_repos:
                ws = wb.create_sheet("Actions Runs")
                runs_headers = [
                    "Repository",
                    "Total Runs (90d)",
                    "Success",
                    "Failure",
                    "Cancelled",
                ]
                _style_header_row(ws, runs_headers)
                for idx, repo in enumerate(runs_repos, 2):
                    ars = repo.actions_run_summary
                    ws.cell(row=idx, column=1, value=repo.name)
                    ws.cell(row=idx, column=2, value=ars.total_runs_90d)
                    ws.cell(row=idx, column=3, value=ars.by_conclusion.get("success", 0))
                    ws.cell(row=idx, column=4, value=ars.by_conclusion.get("failure", 0))
                    ws.cell(row=idx, column=5, value=ars.by_conclusion.get("cancelled", 0))

        # Enterprise sheets (only when enterprise data present)
        if inventory.enterprise is not None:
            ent = inventory.enterprise

            # Enterprise overview sheet
            ws = wb.create_sheet("Enterprise")
            ent_headers = ["Property", "Value"]
            _style_header_row(ws, ent_headers)
            ent_rows: list[tuple[str, str | int | float]] = [
                ("Name", ent.name),
                ("Slug", ent.slug),
                ("Members", ent.members_count),
                ("Admins", ent.admins_count),
                ("Outside Collaborators", ent.outside_collaborators_count),
                ("SAML Enabled", _bool_na(ent.saml.enabled if ent.saml else None)),
                (
                    "IP Allow List Enabled",
                    _bool_na(ent.ip_allow_list.enabled if ent.ip_allow_list else None),
                ),
                (
                    "IP Allow List Entries",
                    ent.ip_allow_list.entries_count if ent.ip_allow_list else 0,
                ),
                (
                    "Verified Domains",
                    ", ".join(ent.verified_domains) if ent.verified_domains else "",
                ),
            ]
            if ent.billing is not None:
                ent_rows.extend(
                    [
                        ("Total Licenses", ent.billing.total_licenses),
                        ("Used Licenses", ent.billing.used_licenses),
                        ("Storage Usage (GB)", ent.billing.storage_usage_gb),
                        ("Storage Quota (GB)", ent.billing.storage_quota_gb),
                        ("Bandwidth Usage (GB)", ent.billing.bandwidth_usage_gb),
                        ("Bandwidth Quota (GB)", ent.billing.bandwidth_quota_gb),
                    ]
                )
            if ent.policies is not None:
                ent_rows.extend(
                    [
                        ("2FA Required", ent.policies.two_factor_required or "n/a"),
                        (
                            "Default Repo Permission",
                            ent.policies.default_repository_permission or "n/a",
                        ),
                    ]
                )
            for idx, (prop, value) in enumerate(ent_rows, 2):
                ws.cell(row=idx, column=1, value=prop)
                ws.cell(row=idx, column=2, value=value)

            # Enterprise Teams sheet
            if ent.enterprise_teams:
                ws = wb.create_sheet("Enterprise Teams")
                teams_headers = ["Name", "Slug", "Members", "Orgs"]
                _style_header_row(ws, teams_headers)
                for idx, team in enumerate(ent.enterprise_teams, 2):
                    ws.cell(row=idx, column=1, value=team.name)
                    ws.cell(row=idx, column=2, value=team.slug)
                    ws.cell(row=idx, column=3, value=team.member_count)
                    ws.cell(row=idx, column=4, value=team.org_count)

        # Apply comprehensive formatting to all sheets
        _apply_formatting(wb)

        wb.save(str(output_path))
