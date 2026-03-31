"""Tests for security detail sections in HTML report and Excel workbook.

Covers:
- HTML security detail section rendered when security detail data is present
- HTML security detail section absent when security_detail is None
- Excel sheets (Dependabot Alerts, Code Scanning Alerts, Secret Scanning Alerts)
  added when security detail present
- Excel security detail sheets absent when security_detail is None
- Data integrity: alert data in sheets
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from gh_audit.models.inventory import Inventory, InventoryMetadata, InventorySummary
from gh_audit.models.repository import RepositoryInventoryItem
from gh_audit.models.security_detail import (
    CodeScanningAlertInfo,
    CodeScanningSetup,
    DependabotAlertInfo,
    SBOMSummary,
    SecretScanningAlertInfo,
    SecurityDetail,
)
from gh_audit.models.user import OrgMemberSummary
from gh_audit.services.excel_export import ExcelExportService
from gh_audit.services.reporting import ReportService


# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------


def _metadata(**kw) -> InventoryMetadata:
    defaults = {
        "schema_version": "2.0",
        "generated_at": datetime(2026, 3, 27, 12, 0, 0, tzinfo=timezone.utc),
        "tool_version": "0.1.0",
        "organization": "test-org",
        "auth_method": "pat",
        "scan_profile": "standard",
    }
    defaults.update(kw)
    return InventoryMetadata(**defaults)


def _make_inventory(**kw) -> Inventory:
    defaults: dict = {
        "metadata": _metadata(),
        "summary": InventorySummary(total_repos=1),
        "repositories": [],
        "users": OrgMemberSummary(total=1),
    }
    defaults.update(kw)
    return Inventory(**defaults)


def _make_repo_with_security_detail() -> RepositoryInventoryItem:
    return RepositoryInventoryItem(
        name="my-repo",
        full_name="test-org/my-repo",
        visibility="private",
        security_detail=SecurityDetail(
            dependabot_alerts=[
                DependabotAlertInfo(
                    severity="critical",
                    package_name="express",
                    manifest_path="package.json",
                    state="open",
                    ghsa_id="GHSA-xxxx",
                    cve_id="CVE-2026-0001",
                    fixed_version="4.18.3",
                ),
                DependabotAlertInfo(
                    severity="high",
                    package_name="lodash",
                    manifest_path="package-lock.json",
                    state="fixed",
                ),
            ],
            code_scanning_alerts=[
                CodeScanningAlertInfo(
                    rule_id="js/xss",
                    severity="error",
                    security_severity="high",
                    tool_name="CodeQL",
                    state="open",
                ),
            ],
            secret_scanning_alerts=[
                SecretScanningAlertInfo(
                    secret_type="github_personal_access_token",
                    secret_type_display_name="GitHub PAT",
                    state="open",
                    push_protection_bypassed=True,
                ),
            ],
            sbom_summary=SBOMSummary(
                dependency_count=42,
                package_managers=["npm", "pip"],
            ),
            code_scanning_setup=CodeScanningSetup(
                default_setup_enabled=True,
                languages=["python", "javascript"],
            ),
            security_configuration_name="org-default",
        ),
    )


def _make_inventory_with_security_detail(**kw) -> Inventory:
    repo = _make_repo_with_security_detail()
    return _make_inventory(
        repositories=[repo],
        **kw,
    )


# ---------------------------------------------------------------------------
# HTML report — security detail section
# ---------------------------------------------------------------------------


class TestHTMLSecurityDetail:
    def test_with_security_detail(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Security Detail" in content

    def test_without_security_detail(self, tmp_path: Path) -> None:
        inv = _make_inventory()  # security_detail=None by default
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        # Security Detail section should not render
        assert "Security Detail" not in content

    def test_dependabot_alerts_table_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Dependabot Alerts (2)" in content
        assert "express" in content
        assert "lodash" in content
        assert "CVE-2026-0001" in content

    def test_code_scanning_alerts_table_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Code Scanning Alerts (1)" in content
        assert "js/xss" in content
        assert "CodeQL" in content

    def test_secret_scanning_alerts_table_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Secret Scanning Alerts (1)" in content
        assert "GitHub PAT" in content

    def test_summary_cards_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Dependabot Alerts" in content
        assert "Code Scanning Alerts" in content
        assert "Secret Scanning Alerts" in content
        assert "Repos with SBOM" in content

    def test_severity_badges_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "critical" in content  # severity badge text
        assert "high" in content

    def test_security_detail_absent_does_not_break_report(self, tmp_path: Path) -> None:
        inv = _make_inventory()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        assert output.exists()
        content = output.read_text()
        assert "<!DOCTYPE html" in content or "<html" in content


# ---------------------------------------------------------------------------
# Excel workbook — security detail sheets
# ---------------------------------------------------------------------------


class TestExcelSecurityDetail:
    def test_with_security_detail_has_sheets(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Dependabot Alerts" in wb.sheetnames
        assert "Code Scanning Alerts" in wb.sheetnames
        assert "Secret Scanning Alerts" in wb.sheetnames

    def test_without_security_detail_no_sheets(self, tmp_path: Path) -> None:
        inv = _make_inventory()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Dependabot Alerts" not in wb.sheetnames
        assert "Code Scanning Alerts" not in wb.sheetnames
        assert "Secret Scanning Alerts" not in wb.sheetnames

    def test_dependabot_alerts_sheet_data(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Dependabot Alerts"]
        # First alert
        assert ws["A2"].value == "my-repo"
        assert ws["B2"].value == "critical"
        assert ws["C2"].value == "express"
        assert ws["D2"].value == "package.json"
        assert ws["E2"].value == "open"
        assert ws["F2"].value == "GHSA-xxxx"
        assert ws["G2"].value == "CVE-2026-0001"
        assert ws["H2"].value == "4.18.3"

    def test_dependabot_alerts_second_row(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Dependabot Alerts"]
        # Second alert
        assert ws["A3"].value == "my-repo"
        assert ws["B3"].value == "high"
        assert ws["C3"].value == "lodash"
        assert ws["E3"].value == "fixed"

    def test_code_scanning_alerts_sheet_data(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Code Scanning Alerts"]
        assert ws["A2"].value == "my-repo"
        assert ws["B2"].value == "js/xss"
        assert ws["C2"].value == "error"
        assert ws["D2"].value == "high"
        assert ws["E2"].value == "CodeQL"
        assert ws["F2"].value == "open"

    def test_secret_scanning_alerts_sheet_data(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Secret Scanning Alerts"]
        assert ws["A2"].value == "my-repo"
        assert ws["B2"].value == "GitHub PAT"  # display name used
        assert ws["C2"].value == "open"
        assert ws["D2"].value in ("", None)  # resolution is None -> empty or None in openpyxl
        assert ws["E2"].value == "Yes"  # push_protection_bypassed=True

    def test_without_security_detail_still_has_core_sheets(self, tmp_path: Path) -> None:
        inv = _make_inventory()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        for sheet in ["Summary", "Repositories", "Warnings"]:
            assert sheet in wb.sheetnames

    def test_with_security_detail_still_has_core_sheets(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        for sheet in [
            "Summary",
            "Repositories",
            "Actions",
            "Security",
            "Issues",
            "Packages",
            "Projects",
            "Users",
            "Large Files",
            "Warnings",
        ]:
            assert sheet in wb.sheetnames

    def test_security_detail_sheets_come_after_warnings(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_security_detail()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        names = wb.sheetnames
        warnings_idx = names.index("Warnings")
        dep_idx = names.index("Dependabot Alerts")
        assert dep_idx > warnings_idx

    def test_empty_security_detail_creates_sheets_with_headers_only(self, tmp_path: Path) -> None:
        """Repo with empty SecurityDetail (scanned, found nothing) still creates sheets."""
        repo = RepositoryInventoryItem(
            name="empty-repo",
            full_name="test-org/empty-repo",
            visibility="private",
            security_detail=SecurityDetail(),
        )
        inv = _make_inventory(repositories=[repo])
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Dependabot Alerts" in wb.sheetnames
        assert "Code Scanning Alerts" in wb.sheetnames
        assert "Secret Scanning Alerts" in wb.sheetnames
        # Should have only header row
        ws = wb["Dependabot Alerts"]
        assert ws["A1"].value == "Repository"
        assert ws["A2"].value is None
