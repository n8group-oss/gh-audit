"""Unit tests for enterprise report sections (HTML + Excel)."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from gh_audit.models.enterprise import (
    EnterpriseBilling,
    EnterpriseInventory,
    EnterpriseIPAllowList,
    EnterprisePolicies,
    EnterpriseSAML,
    EnterpriseTeamInfo,
)
from gh_audit.models.inventory import Inventory, InventoryMetadata, InventorySummary
from gh_audit.models.repository import RepositoryInventoryItem
from gh_audit.models.user import OrgMemberSummary
from gh_audit.services.excel_export import ExcelExportService
from gh_audit.services.reporting import ReportService


def _metadata() -> InventoryMetadata:
    return InventoryMetadata(
        schema_version="2.0",
        generated_at=datetime(2026, 3, 29, tzinfo=timezone.utc),
        tool_version="0.1.0",
        organization="testorg",
        auth_method="pat",
        scan_profile="total",
        enterprise_slug="acme",
        active_categories=["enterprise"],
    )


def _make_inventory(*, with_enterprise: bool = False) -> Inventory:
    repo = RepositoryInventoryItem(name="repo-a", full_name="testorg/repo-a", visibility="public")

    enterprise = None
    if with_enterprise:
        enterprise = EnterpriseInventory(
            name="Acme Corp",
            slug="acme",
            billing=EnterpriseBilling(total_licenses=1000, used_licenses=800),
            policies=EnterprisePolicies(two_factor_required="enabled"),
            saml=EnterpriseSAML(enabled=True, issuer="https://idp.example.com"),
            ip_allow_list=EnterpriseIPAllowList(enabled=True, entries_count=5),
            verified_domains=["acme.com", "acme.io"],
            enterprise_teams=[
                EnterpriseTeamInfo(name="Engineering", slug="eng", member_count=50, org_count=3),
            ],
            members_count=800,
            admins_count=10,
            outside_collaborators_count=20,
        )

    return Inventory(
        metadata=_metadata(),
        summary=InventorySummary(total_repos=1),
        repositories=[repo],
        users=OrgMemberSummary(total=1, admins=0, members=1),
        enterprise=enterprise,
    )


class TestHTMLEnterprise:
    def test_enterprise_section_present(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_enterprise=True)
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Enterprise" in content
        assert "Acme Corp" in content
        assert "SAML" in content
        assert "Billing" in content

    def test_enterprise_section_absent(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_enterprise=False)
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Enterprise</h2>" not in content


class TestExcelEnterprise:
    def test_enterprise_sheet(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_enterprise=True)
        output = tmp_path / "report.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Enterprise" in wb.sheetnames
        ws = wb["Enterprise"]
        assert ws["A1"].value == "Property"
        # Check key data rows exist by scanning column A
        values = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(2, 25)
            if ws.cell(row=r, column=1).value
        }
        assert values["Name"] == "Acme Corp"
        assert values["Total Licenses"] == 1000
        assert values["Used Licenses"] == 800
        assert values["Members"] == 800
        assert values["Admins"] == 10

    def test_enterprise_teams_sheet(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_enterprise=True)
        output = tmp_path / "report.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Enterprise Teams" in wb.sheetnames
        ws = wb["Enterprise Teams"]
        assert ws["A1"].value == "Name"
        assert ws["A2"].value == "Engineering"
        assert ws["C2"].value == 50

    def test_no_enterprise_sheets_without_data(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_enterprise=False)
        output = tmp_path / "report.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Enterprise" not in wb.sheetnames
        assert "Enterprise Teams" not in wb.sheetnames
