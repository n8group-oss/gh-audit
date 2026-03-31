"""Tests for gh_audit.services.excel_export — Excel workbook generation.

Contract enforced:
- Exactly 10 sheets with defined names
- Header row frozen and styled on all data sheets
- Security n/a semantics for None values
- Large Files, Issues, Projects, Warnings populated correctly
- Parent directories created automatically
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from gh_audit.models.actions import ActionsInfo, WorkflowInfo
from gh_audit.models.inventory import Inventory, InventoryMetadata, InventorySummary
from gh_audit.models.packages import PackageInfo
from gh_audit.models.projects import ProjectInfo
from gh_audit.models.repository import LargeFileInfo, LargeFileScan, RepositoryInventoryItem
from gh_audit.models.security import SecurityInfo
from gh_audit.models.user import OrgMemberSummary
from gh_audit.services.excel_export import ExcelExportService


# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------

REQUIRED_SHEETS = [
    "Summary",
    "Repositories",
    "Actions",
    "Security",
    "Issues",
    "Packages",
    "Projects",
    "Users",
    "Large Files",
    "Warnings",
]


def _metadata(**kw) -> InventoryMetadata:
    defaults = {
        "schema_version": "1.0",
        "generated_at": datetime(2026, 3, 27, 12, 0, 0, tzinfo=timezone.utc),
        "tool_version": "0.1.0",
        "organization": "test-org",
        "auth_method": "pat",
        "scan_profile": "standard",
    }
    defaults.update(kw)
    return InventoryMetadata(**defaults)


def _minimal_repo(
    name: str = "repo-a",
    visibility: str = "private",
    *,
    security: SecurityInfo | None = None,
    large_files: list[LargeFileInfo] | None = None,
    warnings: list[str] | None = None,
    actions: ActionsInfo | None = None,
) -> RepositoryInventoryItem:
    lfs = LargeFileScan(
        enabled=True,
        completed=True,
        files=large_files or [],
    )
    return RepositoryInventoryItem(
        name=name,
        full_name=f"test-org/{name}",
        visibility=visibility,
        security=security or SecurityInfo(),
        large_file_scan=lfs,
        actions=actions or ActionsInfo(),
        warnings=warnings or [],
    )


def _minimal_inventory(**kw) -> Inventory:
    defaults: dict = {
        "metadata": _metadata(),
        "summary": InventorySummary(total_repos=1, private_repos=1),
        "repositories": [_minimal_repo()],
        "users": OrgMemberSummary(total=5, admins=1, members=4),
    }
    defaults.update(kw)
    return Inventory(**defaults)


def _generate(inventory: Inventory, tmp_path: Path) -> Path:
    output = tmp_path / "report.xlsx"
    ExcelExportService.generate(inventory, output)
    return output


# ---------------------------------------------------------------------------
# Sheet names contract
# ---------------------------------------------------------------------------


class TestSheetNames:
    def test_all_ten_sheets_present(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        for name in REQUIRED_SHEETS:
            assert name in wb.sheetnames, f"Missing sheet: {name}"

    def test_exactly_ten_sheets(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        assert len(wb.sheetnames) == 10

    def test_issues_sheet_present(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        assert "Issues" in wb.sheetnames

    def test_projects_sheet_present(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        assert "Projects" in wb.sheetnames

    def test_security_sheet_present(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        assert "Security" in wb.sheetnames


# ---------------------------------------------------------------------------
# Security sheet — n/a semantics
# ---------------------------------------------------------------------------


class TestSecuritySheet:
    def test_none_dependabot_renders_na(self, tmp_path: Path) -> None:
        repo = _minimal_repo(
            security=SecurityInfo(
                dependabot_enabled=None,
                dependabot_alerts_open=None,
            )
        )
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Security"]
        # Row 2 is first data row; column B is dependabot_enabled
        assert ws["B2"].value == "n/a"

    def test_true_dependabot_renders_yes(self, tmp_path: Path) -> None:
        repo = _minimal_repo(
            security=SecurityInfo(dependabot_enabled=True, dependabot_alerts_open=3)
        )
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Security"]
        assert ws["B2"].value == "Yes"

    def test_false_dependabot_renders_no(self, tmp_path: Path) -> None:
        repo = _minimal_repo(
            security=SecurityInfo(dependabot_enabled=False, dependabot_alerts_open=None)
        )
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Security"]
        assert ws["B2"].value == "No"

    def test_security_b2_is_yes_no_or_na(self, tmp_path: Path) -> None:
        """Contractual assertion from task spec."""
        repo = _minimal_repo(security=SecurityInfo())
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        assert wb["Security"]["B2"].value in ("Yes", "No", "n/a")

    def test_none_alert_count_renders_na(self, tmp_path: Path) -> None:
        repo = _minimal_repo(
            security=SecurityInfo(
                dependabot_enabled=True,
                dependabot_alerts_open=None,
            )
        )
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Security"]
        # Column C is dependabot_alerts_open
        assert ws["C2"].value == "n/a"

    def test_known_alert_count_renders_number(self, tmp_path: Path) -> None:
        repo = _minimal_repo(
            security=SecurityInfo(
                dependabot_enabled=True,
                dependabot_alerts_open=7,
                counts_exact=True,
            )
        )
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Security"]
        assert ws["C2"].value == 7

    def test_code_scanning_none_renders_na(self, tmp_path: Path) -> None:
        repo = _minimal_repo(
            security=SecurityInfo(code_scanning_enabled=None, code_scanning_alerts_open=None)
        )
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Security"]
        # Column D is code_scanning_enabled
        assert ws["D2"].value == "n/a"

    def test_secret_scanning_none_renders_na(self, tmp_path: Path) -> None:
        repo = _minimal_repo(
            security=SecurityInfo(secret_scanning_enabled=None, secret_scanning_alerts_open=None)
        )
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Security"]
        # Column F is secret_scanning_enabled
        assert ws["F2"].value == "n/a"

    def test_security_has_repo_name_in_column_a(self, tmp_path: Path) -> None:
        repo = _minimal_repo("my-secure-repo")
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Security"]
        assert ws["A2"].value == "my-secure-repo"


# ---------------------------------------------------------------------------
# Repositories sheet
# ---------------------------------------------------------------------------


class TestRepositoriesSheet:
    def test_repositories_has_data_row(self, tmp_path: Path) -> None:
        inv = _minimal_inventory()
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Repositories"]
        # Row 1 = header, row 2 = first data row
        assert ws.max_row >= 2

    def test_repo_name_in_first_data_row(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(repositories=[_minimal_repo("my-repo")])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Repositories"]
        assert ws["A2"].value == "my-repo"

    def test_visibility_in_repositories(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(repositories=[_minimal_repo("pub-repo", "public")])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Repositories"]
        # Column B is visibility
        assert ws["B2"].value == "public"

    def test_multiple_repos_produce_multiple_rows(self, tmp_path: Path) -> None:
        repos = [
            _minimal_repo("alpha"),
            _minimal_repo("beta"),
            _minimal_repo("gamma"),
        ]
        inv = _minimal_inventory(
            repositories=repos,
            summary=InventorySummary(total_repos=3, private_repos=3),
        )
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Repositories"]
        assert ws.max_row == 4  # header + 3 data rows


# ---------------------------------------------------------------------------
# Actions sheet
# ---------------------------------------------------------------------------


class TestActionsSheet:
    def test_actions_sheet_has_header(self, tmp_path: Path) -> None:
        inv = _minimal_inventory()
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Actions"]
        assert ws.max_row >= 1

    def test_workflow_appears_in_actions(self, tmp_path: Path) -> None:
        actions = ActionsInfo(
            has_workflows=True,
            workflow_count=1,
            workflows=[
                WorkflowInfo(
                    name="CI",
                    path=".github/workflows/ci.yml",
                    state="active",
                )
            ],
        )
        repo = _minimal_repo("wf-repo", actions=actions)
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Actions"]
        # Row 2 = first workflow row
        assert ws["A2"].value == "wf-repo"
        assert ws["B2"].value == "CI"

    def test_no_workflows_produces_only_header(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(repositories=[_minimal_repo(actions=ActionsInfo())])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Actions"]
        # max_row is 2 because Excel tables require min 2 rows (header + blank)
        assert ws.max_row <= 2


# ---------------------------------------------------------------------------
# Issues sheet
# ---------------------------------------------------------------------------


class TestIssuesSheet:
    def test_issues_sheet_has_header_row(self, tmp_path: Path) -> None:
        inv = _minimal_inventory()
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Issues"]
        assert ws["A1"].value is not None

    def test_issues_data_row_present(self, tmp_path: Path) -> None:
        repo = RepositoryInventoryItem(
            name="issue-repo",
            full_name="test-org/issue-repo",
            visibility="private",
            issue_count_open=5,
            issue_count_closed=10,
        )
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Issues"]
        assert ws["A2"].value == "issue-repo"
        assert ws["B2"].value == 5
        assert ws["C2"].value == 10

    def test_issues_header_contains_expected_columns(self, tmp_path: Path) -> None:
        inv = _minimal_inventory()
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Issues"]
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        assert any("repo" in str(h).lower() for h in headers)
        assert any("open" in str(h).lower() for h in headers)


# ---------------------------------------------------------------------------
# Packages sheet
# ---------------------------------------------------------------------------


class TestPackagesSheet:
    def test_packages_header_row_exists(self, tmp_path: Path) -> None:
        inv = _minimal_inventory()
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Packages"]
        assert ws["A1"].value is not None

    def test_package_data_appears(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(
            packages=[PackageInfo(name="my-pkg", package_type="npm", visibility="public")]
        )
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Packages"]
        assert ws["A2"].value == "my-pkg"
        assert ws["B2"].value == "npm"

    def test_no_packages_produces_only_header(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(packages=[])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Packages"]
        assert ws.max_row <= 2


# ---------------------------------------------------------------------------
# Projects sheet
# ---------------------------------------------------------------------------


class TestProjectsSheet:
    def test_projects_has_data_when_projects_present(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(
            projects=[ProjectInfo(title="Roadmap", item_count=12, closed=False)]
        )
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Projects"]
        assert ws["A2"].value == "Roadmap"
        assert ws["B2"].value == 12

    def test_closed_project_renders_yes(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(
            projects=[ProjectInfo(title="Done Project", item_count=5, closed=True)]
        )
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Projects"]
        assert ws["C2"].value == "Yes"

    def test_open_project_renders_no(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(
            projects=[ProjectInfo(title="Open Project", item_count=3, closed=False)]
        )
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Projects"]
        assert ws["C2"].value == "No"

    def test_no_projects_produces_only_header(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(projects=[])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Projects"]
        assert ws.max_row <= 2


# ---------------------------------------------------------------------------
# Users sheet
# ---------------------------------------------------------------------------


class TestUsersSheet:
    def test_users_sheet_has_data(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(
            users=OrgMemberSummary(total=10, admins=2, members=6, outside_collaborators=2)
        )
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Users"]
        assert ws.max_row >= 2

    def test_admin_count_present(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(users=OrgMemberSummary(total=10, admins=3, members=7))
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Users"]
        # Find admin row
        values = [
            (ws.cell(row, 1).value, ws.cell(row, 2).value) for row in range(2, ws.max_row + 1)
        ]
        admin_row = [v for v in values if v[0] and "admin" in str(v[0]).lower()]
        assert admin_row, "Admin row not found"
        assert admin_row[0][1] == 3

    def test_total_row_present(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(users=OrgMemberSummary(total=15, admins=2, members=13))
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Users"]
        values = [
            (ws.cell(row, 1).value, ws.cell(row, 2).value) for row in range(2, ws.max_row + 1)
        ]
        total_row = [v for v in values if v[0] and "total" in str(v[0]).lower()]
        assert total_row, "Total row not found"
        assert total_row[0][1] == 15


# ---------------------------------------------------------------------------
# Large Files sheet
# ---------------------------------------------------------------------------


class TestLargeFilesSheet:
    def test_large_files_sheet_has_header(self, tmp_path: Path) -> None:
        inv = _minimal_inventory()
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Large Files"]
        assert ws["A1"].value is not None

    def test_large_file_entry_appears(self, tmp_path: Path) -> None:
        large_files = [LargeFileInfo(path="data/dump.bin", size_bytes=200 * 1024 * 1024)]
        repo = _minimal_repo("big-repo", large_files=large_files)
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Large Files"]
        assert ws["A2"].value == "big-repo"
        assert ws["B2"].value == "data/dump.bin"
        # size_mb should be ~200.0
        assert ws["C2"].value == pytest.approx(200.0, abs=0.1)

    def test_no_large_files_produces_only_header(self, tmp_path: Path) -> None:
        repo = _minimal_repo(large_files=[])
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Large Files"]
        assert ws.max_row <= 2

    def test_multiple_large_files_produce_multiple_rows(self, tmp_path: Path) -> None:
        large_files = [
            LargeFileInfo(path="a.bin", size_bytes=150 * 1024 * 1024),
            LargeFileInfo(path="b.bin", size_bytes=250 * 1024 * 1024),
        ]
        repo = _minimal_repo("multi-big", large_files=large_files)
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Large Files"]
        assert ws.max_row == 3  # header + 2 files


# ---------------------------------------------------------------------------
# Warnings sheet
# ---------------------------------------------------------------------------


class TestWarningsSheet:
    def test_warnings_sheet_has_header(self, tmp_path: Path) -> None:
        inv = _minimal_inventory()
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Warnings"]
        assert ws["A1"].value is not None

    def test_repo_warning_appears(self, tmp_path: Path) -> None:
        repo = _minimal_repo("warn-repo", warnings=["rate limit hit during scan"])
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Warnings"]
        assert ws["A2"].value == "warn-repo"
        assert ws["B2"].value == "rate limit hit during scan"

    def test_scan_level_warning_appears(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(
            metadata=_metadata(scan_warnings=["partial scan: only 50 of 200 repos processed"])
        )
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Warnings"]
        # Scan-level warnings use "Scan" as the repo name
        warning_sources = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
        warning_texts = [ws.cell(row, 2).value for row in range(2, ws.max_row + 1)]
        assert "Scan" in warning_sources
        assert any("partial scan" in str(t) for t in warning_texts)

    def test_no_warnings_produces_only_header(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(metadata=_metadata(scan_warnings=[]))
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Warnings"]
        assert ws.max_row <= 2

    def test_multiple_repo_warnings(self, tmp_path: Path) -> None:
        repo = _minimal_repo("repo-w", warnings=["warn1", "warn2"])
        inv = _minimal_inventory(repositories=[repo])
        wb = load_workbook(_generate(inv, tmp_path))
        ws = wb["Warnings"]
        # 1 header + 2 warnings
        assert ws.max_row == 3


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------


class TestSummarySheet:
    def _all_values(self, ws):
        """Collect all cell values across columns A-C (3-column summary layout)."""
        return [
            ws.cell(row, col).value
            for row in range(1, ws.max_row + 1)
            for col in range(1, 4)
        ]

    def test_summary_has_org_name(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(metadata=_metadata(organization="acme-corp"))
        wb = load_workbook(_generate(inv, tmp_path))
        assert "acme-corp" in self._all_values(wb["Summary"])

    def test_summary_has_total_repos(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(summary=InventorySummary(total_repos=42, private_repos=42))
        wb = load_workbook(_generate(inv, tmp_path))
        assert 42 in self._all_values(wb["Summary"])

    def test_summary_has_scan_profile(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(metadata=_metadata(scan_profile="full"))
        wb = load_workbook(_generate(inv, tmp_path))
        assert "full" in self._all_values(wb["Summary"])

    def test_summary_has_tool_version(self, tmp_path: Path) -> None:
        inv = _minimal_inventory(metadata=_metadata(tool_version="1.2.3"))
        wb = load_workbook(_generate(inv, tmp_path))
        assert "1.2.3" in self._all_values(wb["Summary"])


# ---------------------------------------------------------------------------
# Formatting — frozen header rows
# ---------------------------------------------------------------------------


class TestFrozenHeaders:
    def _check_frozen(self, ws) -> bool:
        """Return True if the sheet has a freeze at row 2 (i.e. header is frozen).

        ``freeze_panes`` is stored as a cell address string like ``"A2"``.
        We check that it ends with ``"2"`` (row 2 frozen = header is frozen).
        """
        fp = ws.freeze_panes
        if fp is None:
            return False
        # Cell address string e.g. "A2" — extract row number
        from openpyxl.utils.cell import coordinate_to_tuple

        row, _col = coordinate_to_tuple(fp)
        return row >= 2

    def test_repositories_header_frozen(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        assert self._check_frozen(wb["Repositories"])

    def test_security_header_frozen(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        assert self._check_frozen(wb["Security"])

    def test_actions_header_frozen(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        assert self._check_frozen(wb["Actions"])

    def test_issues_header_frozen(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        assert self._check_frozen(wb["Issues"])

    def test_large_files_header_frozen(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        assert self._check_frozen(wb["Large Files"])

    def test_warnings_header_frozen(self, tmp_path: Path) -> None:
        wb = load_workbook(_generate(_minimal_inventory(), tmp_path))
        assert self._check_frozen(wb["Warnings"])


# ---------------------------------------------------------------------------
# Static method contract
# ---------------------------------------------------------------------------


class TestStaticMethodContract:
    def test_generate_is_static(self) -> None:
        assert isinstance(ExcelExportService.__dict__["generate"], staticmethod)

    def test_generate_returns_none(self, tmp_path: Path) -> None:
        result = ExcelExportService.generate(_minimal_inventory(), tmp_path / "out.xlsx")
        assert result is None

    def test_creates_xlsx_file(self, tmp_path: Path) -> None:
        output = tmp_path / "report.xlsx"
        ExcelExportService.generate(_minimal_inventory(), output)
        assert output.exists()


# ---------------------------------------------------------------------------
# Parent directory creation
# ---------------------------------------------------------------------------


class TestParentDirCreation:
    def test_creates_nested_parent_dirs(self, tmp_path: Path) -> None:
        output = tmp_path / "deep" / "nested" / "report.xlsx"
        ExcelExportService.generate(_minimal_inventory(), output)
        assert output.exists()

    def test_existing_parent_dir_ok(self, tmp_path: Path) -> None:
        output = tmp_path / "report.xlsx"
        ExcelExportService.generate(_minimal_inventory(), output)
        # Second call should overwrite without error
        ExcelExportService.generate(_minimal_inventory(), output)
        assert output.exists()
