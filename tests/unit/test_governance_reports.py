"""Tests for governance sections in HTML report and Excel workbook.

Covers:
- HTML governance section rendered when governance data is present
- HTML governance section absent when governance is None
- Excel sheets (Teams, Org Policies, Org Rulesets) added when governance present
- Excel governance sheets absent when governance is None
- Data integrity: team name in Teams sheet, policy labels in Org Policies sheet
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from gh_audit.models.governance import (
    CustomPropertySchema,
    CustomRoleInfo,
    GovernanceInventory,
    OrgPolicies,
    RulesetDetail,
    TeamInfo,
)
from gh_audit.models.inventory import Inventory, InventoryMetadata, InventorySummary
from gh_audit.models.user import OrgMemberSummary
from gh_audit.services.excel_export import ExcelExportService
from gh_audit.services.reporting import ReportService


# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------


def _metadata(**kw) -> InventoryMetadata:
    defaults = {
        "schema_version": "2.0",
        "generated_at": datetime(2026, 3, 27, 12, 0, 0, tzinfo=timezone.utc),
        "tool_version": "0.1.0",
        "organization": "test-org",
        "auth_method": "pat",
        "scan_profile": "standard",
    }
    defaults.update(kw)
    return InventoryMetadata(**defaults)


def _make_inventory(**kw) -> Inventory:
    defaults: dict = {
        "metadata": _metadata(),
        "summary": InventorySummary(total_repos=1),
        "repositories": [],
        "users": OrgMemberSummary(total=1),
    }
    defaults.update(kw)
    return Inventory(**defaults)


def _make_inventory_with_governance(**kw) -> Inventory:
    gov = GovernanceInventory(
        teams=[
            TeamInfo(
                name="backend",
                slug="backend",
                privacy="closed",
                permission="push",
                member_count=5,
                repo_count=10,
            )
        ],
        org_rulesets=[
            RulesetDetail(
                name="main-protect",
                enforcement="active",
                target="branch",
                source_type="Organization",
            )
        ],
        org_policies=OrgPolicies(
            two_factor_requirement_enabled=True,
            default_repository_permission="read",
        ),
        custom_roles=[CustomRoleInfo(name="reviewer")],
        custom_properties_schema=[CustomPropertySchema(property_name="team", value_type="string")],
        org_secrets_count=3,
        org_variables_count=5,
    )
    return _make_inventory(governance=gov, **kw)


# ---------------------------------------------------------------------------
# HTML report — governance section
# ---------------------------------------------------------------------------


class TestHTMLGovernance:
    def test_with_governance(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Governance" in content
        assert "backend" in content  # team name from fixture

    def test_without_governance(self, tmp_path: Path) -> None:
        inv = _make_inventory()  # governance=None by default
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        # Governance section should not render
        assert "Organization Policies" not in content

    def test_teams_section_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Teams" in content

    def test_policies_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "2FA required" in content

    def test_rulesets_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "main-protect" in content

    def test_org_secrets_count_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Org Secrets" in content

    def test_governance_absent_does_not_break_report(self, tmp_path: Path) -> None:
        """Report with governance=None must still generate a valid HTML file."""
        inv = _make_inventory()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        assert output.exists()
        content = output.read_text()
        assert "<!DOCTYPE html" in content or "<html" in content


# ---------------------------------------------------------------------------
# Excel workbook — governance sheets
# ---------------------------------------------------------------------------


class TestExcelGovernance:
    def test_with_governance_has_sheets(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Teams" in wb.sheetnames
        assert "Org Policies" in wb.sheetnames
        assert "Org Rulesets" in wb.sheetnames

    def test_without_governance_no_sheets(self, tmp_path: Path) -> None:
        inv = _make_inventory()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Teams" not in wb.sheetnames
        assert "Org Policies" not in wb.sheetnames
        assert "Org Rulesets" not in wb.sheetnames

    def test_teams_sheet_data(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert wb["Teams"]["A2"].value == "backend"

    def test_teams_sheet_member_count(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Teams"]
        # Column D = Members
        assert ws["D2"].value == 5

    def test_teams_sheet_repo_count(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Teams"]
        # Column E = Repos
        assert ws["E2"].value == 10

    def test_org_policies_sheet_has_data(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Org Policies"]
        # Header row exists
        assert ws["A1"].value is not None
        # At least one data row
        assert ws.max_row >= 2

    def test_org_policies_2fa_yes(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Org Policies"]
        # Find the 2FA row by scanning column A
        policy_values = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
        assert policy_values.get("2FA required") == "Yes"

    def test_org_rulesets_sheet_data(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_governance()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Org Rulesets"]
        assert ws["A2"].value == "main-protect"
        assert ws["B2"].value == "active"

    def test_without_governance_still_has_core_sheets(self, tmp_path: Path) -> None:
        """Core 10 sheets must always be present regardless of governance."""
        inv = _make_inventory()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        for sheet in ["Summary", "Repositories", "Warnings"]:
            assert sheet in wb.sheetnames

    def test_with_governance_still_has_core_sheets(self, tmp_path: Path) -> None:
        """Adding governance sheets must not displace the core 10 sheets."""
        inv = _make_inventory_with_governance()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        for sheet in [
            "Summary",
            "Repositories",
            "Actions",
            "Security",
            "Issues",
            "Packages",
            "Projects",
            "Users",
            "Large Files",
            "Warnings",
        ]:
            assert sheet in wb.sheetnames

    def test_governance_sheets_come_after_warnings(self, tmp_path: Path) -> None:
        """Teams, Org Policies, Org Rulesets must appear after Warnings in sheet order."""
        inv = _make_inventory_with_governance()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        names = wb.sheetnames
        warnings_idx = names.index("Warnings")
        teams_idx = names.index("Teams")
        assert teams_idx > warnings_idx
