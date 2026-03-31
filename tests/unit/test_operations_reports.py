"""Tests for operations sections in HTML report and Excel workbook.

Covers:
- HTML operations section rendered when operations data is present
- HTML operations section absent when operations is None
- Excel sheets (Runners, Environments, Installed Apps) added when operations present
- Excel operations sheets absent when operations is None
- Data integrity: runner name in Runners sheet, environment data in Environments sheet
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from gh_audit.models.inventory import Inventory, InventoryMetadata, InventorySummary
from gh_audit.models.operations import (
    ActionsPermissions,
    DeployKeyInfo,
    EnvironmentInfo,
    EnvironmentProtection,
    InstalledAppInfo,
    OperationsInventory,
    RunnerGroupInfo,
    RunnerInfo,
    WebhookInfo,
)
from gh_audit.models.repository import RepositoryInventoryItem
from gh_audit.models.user import OrgMemberSummary
from gh_audit.services.excel_export import ExcelExportService
from gh_audit.services.reporting import ReportService


# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------


def _metadata(**kw) -> InventoryMetadata:
    defaults = {
        "schema_version": "2.0",
        "generated_at": datetime(2026, 3, 27, 12, 0, 0, tzinfo=timezone.utc),
        "tool_version": "0.1.0",
        "organization": "test-org",
        "auth_method": "pat",
        "scan_profile": "standard",
    }
    defaults.update(kw)
    return InventoryMetadata(**defaults)


def _make_inventory(**kw) -> Inventory:
    defaults: dict = {
        "metadata": _metadata(),
        "summary": InventorySummary(total_repos=1),
        "repositories": [],
        "users": OrgMemberSummary(total=1),
    }
    defaults.update(kw)
    return Inventory(**defaults)


def _make_repo_with_operations() -> RepositoryInventoryItem:
    return RepositoryInventoryItem(
        name="my-repo",
        full_name="test-org/my-repo",
        visibility="private",
        environments=[
            EnvironmentInfo(
                name="production",
                protection_rules=EnvironmentProtection(
                    wait_timer=30,
                    required_reviewers=2,
                    branch_policy="protected",
                ),
                can_admins_bypass=False,
            ),
            EnvironmentInfo(
                name="staging",
            ),
        ],
        deploy_keys=[
            DeployKeyInfo(title="deploy-key-1", read_only=True, created_at="2025-01-01T00:00:00Z"),
        ],
        repo_webhooks=[
            WebhookInfo(url_domain="ci.example.com", events=["push"], active=True),
        ],
        repo_secrets_count=3,
        repo_variables_count=2,
        actions_permissions=ActionsPermissions(enabled=True, allowed_actions="selected"),
    )


def _make_inventory_with_operations(**kw) -> Inventory:
    ops = OperationsInventory(
        runners=[
            RunnerInfo(
                name="runner-1",
                os="Linux",
                status="online",
                labels=["self-hosted", "linux"],
                busy=False,
                runner_group_name="Default",
            ),
        ],
        runner_groups=[
            RunnerGroupInfo(
                name="Default",
                visibility="all",
                runner_count=1,
            ),
        ],
        installed_apps=[
            InstalledAppInfo(
                app_name="Renovate",
                app_slug="renovate",
                permissions={"issues": "read"},
                events=["push", "pull_request"],
                repository_selection="selected",
            ),
        ],
        org_webhooks=[
            WebhookInfo(
                url_domain="hooks.slack.com",
                events=["push"],
                active=True,
                content_type="json",
            ),
        ],
    )
    repo = _make_repo_with_operations()
    return _make_inventory(
        operations=ops,
        repositories=[repo],
        **kw,
    )


# ---------------------------------------------------------------------------
# HTML report — operations section
# ---------------------------------------------------------------------------


class TestHTMLOperations:
    def test_with_operations(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Operations" in content
        assert "runner-1" in content  # runner name from fixture

    def test_without_operations(self, tmp_path: Path) -> None:
        inv = _make_inventory()  # operations=None by default
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        # Operations section should not render
        assert "Runners" not in content or "Self-Hosted Runners" in content
        # More precise: operations-specific tables should not render
        assert "Runner Groups" not in content

    def test_runners_section_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Runners (1)" in content
        assert "runner-1" in content
        assert "Linux" in content

    def test_runner_groups_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Runner Groups (1)" in content
        assert "Default" in content

    def test_installed_apps_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Installed Apps (1)" in content
        assert "Renovate" in content

    def test_org_webhooks_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Organization Webhooks (1)" in content
        assert "hooks.slack.com" in content

    def test_summary_cards_rendered(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        # Check that the summary cards are present
        assert "Runner Groups" in content
        assert "Installed Apps" in content
        assert "Webhooks" in content

    def test_operations_absent_does_not_break_report(self, tmp_path: Path) -> None:
        """Report with operations=None must still generate a valid HTML file."""
        inv = _make_inventory()
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        assert output.exists()
        content = output.read_text()
        assert "<!DOCTYPE html" in content or "<html" in content


# ---------------------------------------------------------------------------
# Excel workbook — operations sheets
# ---------------------------------------------------------------------------


class TestExcelOperations:
    def test_with_operations_has_sheets(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Runners" in wb.sheetnames
        assert "Environments" in wb.sheetnames
        assert "Installed Apps" in wb.sheetnames

    def test_without_operations_no_sheets(self, tmp_path: Path) -> None:
        inv = _make_inventory()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Runners" not in wb.sheetnames
        assert "Environments" not in wb.sheetnames
        assert "Installed Apps" not in wb.sheetnames

    def test_runners_sheet_data(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Runners"]
        assert ws["A2"].value == "runner-1"
        assert ws["B2"].value == "Linux"
        assert ws["C2"].value == "online"

    def test_runners_sheet_labels(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Runners"]
        assert ws["D2"].value == "self-hosted, linux"

    def test_runners_sheet_group(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Runners"]
        assert ws["E2"].value == "Default"

    def test_environments_sheet_data(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Environments"]
        # First environment: production
        assert ws["A2"].value == "my-repo"
        assert ws["B2"].value == "production"
        assert ws["C2"].value == 30  # wait timer
        assert ws["D2"].value == 2  # required reviewers
        assert ws["E2"].value == "protected"  # branch policy
        assert ws["F2"].value == "No"  # can_admins_bypass = False

    def test_environments_sheet_staging(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Environments"]
        # Second environment: staging (no protection)
        assert ws["A3"].value == "my-repo"
        assert ws["B3"].value == "staging"
        assert ws["C3"].value == 0  # no protection -> 0
        assert ws["D3"].value == 0

    def test_installed_apps_sheet_data(self, tmp_path: Path) -> None:
        inv = _make_inventory_with_operations()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        ws = wb["Installed Apps"]
        assert ws["A2"].value == "Renovate"
        assert ws["B2"].value == "renovate"
        assert ws["C2"].value == "selected"
        assert ws["D2"].value == 2  # events count

    def test_without_operations_still_has_core_sheets(self, tmp_path: Path) -> None:
        """Core 10 sheets must always be present regardless of operations."""
        inv = _make_inventory()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        for sheet in ["Summary", "Repositories", "Warnings"]:
            assert sheet in wb.sheetnames

    def test_with_operations_still_has_core_sheets(self, tmp_path: Path) -> None:
        """Adding operations sheets must not displace the core 10 sheets."""
        inv = _make_inventory_with_operations()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        for sheet in [
            "Summary",
            "Repositories",
            "Actions",
            "Security",
            "Issues",
            "Packages",
            "Projects",
            "Users",
            "Large Files",
            "Warnings",
        ]:
            assert sheet in wb.sheetnames

    def test_operations_sheets_come_after_warnings(self, tmp_path: Path) -> None:
        """Runners, Environments, Installed Apps must appear after Warnings in sheet order."""
        inv = _make_inventory_with_operations()
        output = tmp_path / "test.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        names = wb.sheetnames
        warnings_idx = names.index("Warnings")
        runners_idx = names.index("Runners")
        assert runners_idx > warnings_idx
