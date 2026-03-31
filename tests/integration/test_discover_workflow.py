"""End-to-end integration tests for the gh-audit discovery workflow.

These tests exercise the full discovery pipeline using mocked HTTP clients,
verifying that:
- ScannerConfig profiles control which REST calls are made
- Inventory is populated correctly from mocked API responses
- JSON round-trip serialisation preserves all fields
- HTML report is self-contained (no CDN links) and contains the org name
- Excel workbook has the required 10 sheets
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import AsyncMock

import pytest
from openpyxl import load_workbook
from pydantic import SecretStr

from gh_audit.adapters.base import AlertCountResult
from gh_audit.models.config import ScannerConfig
from gh_audit.models.inventory import Inventory
from gh_audit.services.discovery import DiscoveryService
from gh_audit.services.excel_export import ExcelExportService
from gh_audit.services.reporting import ReportService


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

_ORG = "integration-org"

_WORKFLOW_YAML = """\
name: CI
on: [push]
jobs:
  build:
    runs-on: self-hosted
    steps:
      - uses: actions/checkout@v4
      - uses: actions/setup-python@v5
"""

_EXPECTED_EXCEL_SHEETS = [
    "Summary",
    "Repositories",
    "Actions",
    "Security",
    "Issues",
    "Packages",
    "Projects",
    "Users",
    "Large Files",
    "Warnings",
]


def _make_config(**kwargs) -> ScannerConfig:
    """Build a ScannerConfig with PAT auth, overridable via kwargs."""
    defaults = {
        "token": SecretStr("ghp_integration_test"),
        "organization": _ORG,
    }
    defaults.update(kwargs)
    return ScannerConfig(**defaults)


def _make_graphql_repo(
    name: str = "alpha",
    *,
    org: str = _ORG,
    visibility: str = "PRIVATE",
    default_branch: str = "main",
    disk_usage: int = 2048,
) -> dict:
    """Return a minimal GraphQL repo node."""
    return {
        "name": name,
        "nameWithOwner": f"{org}/{name}",
        "visibility": visibility,
        "isArchived": False,
        "isFork": False,
        "isTemplate": False,
        "primaryLanguage": {"name": "TypeScript"},
        "repositoryTopics": {"nodes": []},
        "diskUsage": disk_usage,
        "defaultBranchRef": {"name": default_branch},
        "description": f"Integration test repo {name}",
        "refs": {"totalCount": 4},
        "openPRs": {"totalCount": 1},
        "closedPRs": {"totalCount": 2},
        "mergedPRs": {"totalCount": 8},
        "openIssues": {"totalCount": 3},
        "closedIssues": {"totalCount": 6},
        "labels": {"nodes": [{"name": "bug", "issues": {"totalCount": 2}}]},
        "branchProtectionRules": {"totalCount": 1},
        "object": None,
    }


def _make_standard_rest() -> AsyncMock:
    """Mock REST client with standard-profile defaults."""
    rest = AsyncMock()
    rest.list_workflows.return_value = [
        {"name": "CI", "path": ".github/workflows/ci.yml", "state": "active"},
    ]
    rest.get_workflow_file.return_value = None  # not called in standard profile
    rest.get_tree.return_value = {"tree": [], "truncated": False}
    rest.count_dependabot_alerts.return_value = AlertCountResult.inaccessible()
    rest.count_code_scanning_alerts.return_value = AlertCountResult.inaccessible()
    rest.count_secret_scanning_alerts.return_value = AlertCountResult.inaccessible()
    rest.get_security_features.return_value = {
        "security_and_analysis": {
            "advanced_security": {"status": "disabled"},
            "dependabot_security_updates": {"status": "enabled"},
            "secret_scanning": {"status": "disabled"},
        }
    }
    rest.list_rulesets.return_value = [{"id": 1, "name": "main-protect"}]
    rest.list_org_members.return_value = [{"login": "alice"}, {"login": "bob"}]
    rest.list_outside_collaborators.return_value = []
    rest.list_packages.return_value = []
    return rest


def _make_deep_rest() -> AsyncMock:
    """Mock REST client for deep profile: large files + workflow contents + alerts."""
    rest = _make_standard_rest()
    # Large-file tree: one huge file (200 MB) and one small file
    rest.get_tree.return_value = {
        "tree": [
            {"type": "blob", "path": "assets/huge.bin", "size": 209_715_200},  # 200 MB
            {"type": "blob", "path": "src/main.ts", "size": 4096},
        ],
        "truncated": False,
    }
    # Workflow content with self-hosted runner and external action
    rest.get_workflow_file.return_value = _WORKFLOW_YAML
    # Alert counts: all accessible, count=3 each
    rest.count_dependabot_alerts.return_value = AlertCountResult.from_count(3)
    rest.count_code_scanning_alerts.return_value = AlertCountResult.from_count(3)
    rest.count_secret_scanning_alerts.return_value = AlertCountResult.from_count(3)
    return rest


def _make_gql(repos: list[dict] | None = None) -> AsyncMock:
    """Mock GraphQL client returning the given repo list."""
    gql = AsyncMock()
    gql.fetch_all_repos.return_value = repos if repos is not None else [_make_graphql_repo()]
    gql.fetch_projects.return_value = []
    return gql


# ---------------------------------------------------------------------------
# Standard profile integration tests
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.asyncio
class TestStandardProfileIntegration:
    """Full pipeline using standard profile (no large files, no deep parse, no counts)."""

    async def test_metadata_scan_profile(self):
        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        assert inventory.metadata.scan_profile == "standard"
        assert inventory.metadata.organization == _ORG

    async def test_security_alert_counts_absent(self):
        """Standard profile must not populate security alert counts."""
        config = _make_config(scan_profile="standard", security_alert_counts=False)
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        repo = inventory.repositories[0]
        assert repo.security.dependabot_alerts_open is None
        assert repo.security.code_scanning_alerts_open is None
        assert repo.security.secret_scanning_alerts_open is None

    async def test_large_file_scan_not_completed(self):
        """Standard profile must not complete a large-file scan."""
        config = _make_config(scan_profile="standard", scan_large_files=False)
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        repo = inventory.repositories[0]
        assert repo.large_file_scan.completed is False
        assert repo.large_file_scan.enabled is False

    async def test_actions_analysis_level_listing(self):
        """Standard profile without workflow-content parsing yields 'listing' level."""
        config = _make_config(scan_profile="standard", scan_workflow_contents=False)
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        repo = inventory.repositories[0]
        assert repo.actions.analysis_level == "listing"
        assert repo.actions.has_workflows is True
        assert repo.actions.workflow_count == 1

    async def test_multi_repo_standard(self):
        """Two repos both processed under standard profile."""
        repos = [_make_graphql_repo("alpha"), _make_graphql_repo("beta")]
        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(repos),
            config=config,
        )
        inventory = await svc.discover()

        assert len(inventory.repositories) == 2
        assert inventory.summary.total_repos == 2
        names = {r.name for r in inventory.repositories}
        assert names == {"alpha", "beta"}

    async def test_security_features_populated(self):
        """REST get_security_features results are mapped into SecurityInfo."""
        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        sec = inventory.repositories[0].security
        assert sec.dependabot_enabled is True
        assert sec.code_scanning_enabled is False
        assert sec.secret_scanning_enabled is False

    async def test_rulesets_populated(self):
        """REST list_rulesets count is reflected in branch_protection."""
        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        bp = inventory.repositories[0].branch_protection
        assert bp.ruleset_count == 1

    async def test_summary_aggregation(self):
        """InventorySummary aggregates repo-level data correctly."""
        repos = [
            _make_graphql_repo("alpha"),
            _make_graphql_repo("beta", visibility="PUBLIC"),
        ]
        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(repos),
            config=config,
        )
        inventory = await svc.discover()

        summary = inventory.summary
        assert summary.total_repos == 2
        assert summary.public_repos == 1
        assert summary.private_repos == 1
        assert summary.repos_with_workflows == 2


# ---------------------------------------------------------------------------
# Deep profile integration tests
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.asyncio
class TestDeepProfileIntegration:
    """Full pipeline using deep profile (large files + workflow parse + alert counts)."""

    async def test_large_file_scan_completed(self):
        config = _make_config(
            scan_profile="deep",
            scan_large_files=True,
            scan_workflow_contents=True,
            security_alert_counts=True,
        )
        svc = DiscoveryService(
            rest_client=_make_deep_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        repo = inventory.repositories[0]
        assert repo.large_file_scan.completed is True
        assert repo.large_file_scan.enabled is True

    async def test_large_file_detected(self):
        """The 200 MB file must be in the large_file_scan.files list."""
        config = _make_config(scan_large_files=True)
        svc = DiscoveryService(
            rest_client=_make_deep_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        files = inventory.repositories[0].large_file_scan.files
        assert len(files) == 1
        assert files[0].path == "assets/huge.bin"
        assert files[0].size_bytes == 209_715_200

    async def test_workflow_analysis_level_parsed(self):
        """With scan_workflow_contents=True, analysis_level must be 'parsed'."""
        config = _make_config(
            scan_workflow_contents=True,
        )
        svc = DiscoveryService(
            rest_client=_make_deep_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        repo = inventory.repositories[0]
        assert repo.actions.analysis_level == "parsed"

    async def test_self_hosted_runner_detected(self):
        """runs-on: self-hosted in YAML must flip uses_self_hosted_runners."""
        config = _make_config(scan_workflow_contents=True)
        svc = DiscoveryService(
            rest_client=_make_deep_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        repo = inventory.repositories[0]
        assert repo.actions.uses_self_hosted_runners is True

    async def test_external_actions_extracted(self):
        """actions/checkout@v4 and actions/setup-python@v5 must appear in actions_used."""
        config = _make_config(scan_workflow_contents=True)
        svc = DiscoveryService(
            rest_client=_make_deep_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        used = inventory.repositories[0].actions.actions_used
        assert "actions/checkout@v4" in used
        assert "actions/setup-python@v5" in used

    async def test_security_alert_counts_populated(self):
        """With security_alert_counts=True all three counts must be set."""
        config = _make_config(security_alert_counts=True)
        svc = DiscoveryService(
            rest_client=_make_deep_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        sec = inventory.repositories[0].security
        assert sec.dependabot_alerts_open == 3
        assert sec.code_scanning_alerts_open == 3
        assert sec.secret_scanning_alerts_open == 3

    async def test_counts_exact_true_when_all_accessible(self):
        """counts_exact should be True when all three alert endpoints succeed."""
        config = _make_config(security_alert_counts=True)
        svc = DiscoveryService(
            rest_client=_make_deep_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        assert inventory.repositories[0].security.counts_exact is True

    async def test_summary_large_files_aggregation(self):
        """InventorySummary.repos_with_large_files must count repos with files."""
        config = _make_config(scan_large_files=True)
        svc = DiscoveryService(
            rest_client=_make_deep_rest(),
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        assert inventory.summary.repos_with_large_files == 1


# ---------------------------------------------------------------------------
# Graceful degradation tests
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.asyncio
class TestGracefulDegradation:
    """Discovery must complete even when optional REST calls fail/return None."""

    async def test_forbidden_rulesets_adds_warning(self):
        """When list_rulesets returns None (forbidden), a warning is added."""
        rest = _make_standard_rest()
        rest.list_rulesets.return_value = None  # simulate 403

        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=rest,
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        repo = inventory.repositories[0]
        assert repo.branch_protection.ruleset_count is None
        assert any("Rulesets not accessible" in w for w in repo.warnings)

    async def test_security_features_exception_adds_warning(self):
        """When get_security_features raises, a warning is recorded but scan continues."""
        rest = _make_standard_rest()
        rest.get_security_features.side_effect = Exception("network timeout")

        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=rest,
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        # Inventory still populated
        assert len(inventory.repositories) == 1
        repo = inventory.repositories[0]
        assert any("Failed to get security features" in w for w in repo.warnings)
        # Security fields remain unset
        assert repo.security.dependabot_enabled is None

    async def test_workflow_list_exception_adds_warning(self):
        """When list_workflows raises, workflow enrichment degrades gracefully."""
        rest = _make_standard_rest()
        rest.list_workflows.side_effect = Exception("connection reset")

        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=rest,
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        assert len(inventory.repositories) == 1
        repo = inventory.repositories[0]
        assert any("Failed to list workflows" in w for w in repo.warnings)

    async def test_large_file_scan_exception_adds_warning(self):
        """When get_tree raises, large_file_scan stays incomplete but scan continues."""
        rest = _make_standard_rest()
        rest.get_tree.side_effect = Exception("timeout")

        config = _make_config(scan_large_files=True)
        svc = DiscoveryService(
            rest_client=rest,
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        repo = inventory.repositories[0]
        assert repo.large_file_scan.completed is False
        assert any("Failed to scan large files" in w for w in repo.warnings)

    async def test_user_discovery_exception_adds_scan_warning(self):
        """When list_org_members raises, users default to zero counts."""
        rest = _make_standard_rest()
        rest.list_org_members.side_effect = Exception("forbidden")

        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=rest,
            graphql_client=_make_gql(),
            config=config,
        )
        inventory = await svc.discover()

        assert inventory.users.total == 0
        assert any("Failed to discover org members" in w for w in inventory.metadata.scan_warnings)

    async def test_inventory_complete_despite_multiple_failures(self):
        """Multiple simultaneous failures produce warnings but a complete Inventory."""
        rest = _make_standard_rest()
        rest.list_rulesets.return_value = None
        rest.get_security_features.side_effect = Exception("boom")

        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=rest,
            graphql_client=_make_gql(
                [
                    _make_graphql_repo("alpha"),
                    _make_graphql_repo("beta"),
                ]
            ),
            config=config,
        )
        inventory = await svc.discover()

        assert len(inventory.repositories) == 2
        for repo in inventory.repositories:
            assert len(repo.warnings) >= 1  # at least one warning per repo


# ---------------------------------------------------------------------------
# Report generation integration tests
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.asyncio
class TestReportGenerationIntegration:
    """Report generation from a real (mocked) Inventory."""

    async def _run_standard_discovery(self) -> Inventory:
        config = _make_config(scan_profile="standard")
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(
                [
                    _make_graphql_repo("alpha"),
                    _make_graphql_repo("beta", visibility="PUBLIC"),
                ]
            ),
            config=config,
        )
        return await svc.discover()

    async def test_json_roundtrip(self, tmp_path: Path):
        """Inventory serialises to JSON and round-trips back losslessly."""
        inventory = await self._run_standard_discovery()

        json_path = tmp_path / "inventory.json"
        json_path.write_text(inventory.model_dump_json(indent=2), encoding="utf-8")

        loaded = Inventory.model_validate_json(json_path.read_text(encoding="utf-8"))

        assert loaded.metadata.organization == _ORG
        assert loaded.metadata.scan_profile == "standard"
        assert len(loaded.repositories) == 2
        assert loaded.summary.total_repos == 2

    async def test_json_roundtrip_repo_fields_preserved(self, tmp_path: Path):
        """Key repo sub-fields survive the JSON round-trip unchanged."""
        inventory = await self._run_standard_discovery()
        json_path = tmp_path / "inventory.json"
        json_path.write_text(inventory.model_dump_json(), encoding="utf-8")

        loaded = Inventory.model_validate_json(json_path.read_text(encoding="utf-8"))
        repo = loaded.repositories[0]

        assert repo.actions.analysis_level == "listing"
        assert repo.large_file_scan.completed is False
        assert repo.security.dependabot_alerts_open is None

    async def test_html_report_file_created(self, tmp_path: Path):
        """ReportService must create an HTML file at the given path."""
        inventory = await self._run_standard_discovery()
        report_path = tmp_path / "report.html"
        ReportService().generate(inventory, report_path)

        assert report_path.exists()
        assert report_path.stat().st_size > 0

    async def test_html_report_contains_org_name(self, tmp_path: Path):
        """The generated HTML must contain the organisation name."""
        inventory = await self._run_standard_discovery()
        report_path = tmp_path / "report.html"
        ReportService().generate(inventory, report_path)

        content = report_path.read_text(encoding="utf-8")
        assert _ORG in content

    async def test_html_report_no_cdn_links(self, tmp_path: Path):
        """The report must not contain external CDN URLs."""
        inventory = await self._run_standard_discovery()
        report_path = tmp_path / "report.html"
        ReportService().generate(inventory, report_path)

        content = report_path.read_text(encoding="utf-8")
        cdn_patterns = [
            "cdn.jsdelivr.net",
            "cdnjs.cloudflare.com",
            "unpkg.com",
            "fonts.googleapis.com",
            "bootstrapcdn.com",
        ]
        for pattern in cdn_patterns:
            assert pattern not in content, f"Found CDN link: {pattern}"

    async def test_excel_workbook_file_created(self, tmp_path: Path):
        """ExcelExportService must write a valid .xlsx file."""
        inventory = await self._run_standard_discovery()
        xlsx_path = tmp_path / "report.xlsx"
        ExcelExportService.generate(inventory, xlsx_path)

        assert xlsx_path.exists()
        assert xlsx_path.stat().st_size > 0

    async def test_excel_workbook_has_ten_sheets(self, tmp_path: Path):
        """The workbook must contain exactly 10 sheets in the defined order."""
        inventory = await self._run_standard_discovery()
        xlsx_path = tmp_path / "report.xlsx"
        ExcelExportService.generate(inventory, xlsx_path)

        wb = load_workbook(str(xlsx_path), read_only=True)
        assert wb.sheetnames == _EXPECTED_EXCEL_SHEETS
        wb.close()

    async def test_excel_repositories_sheet_has_data(self, tmp_path: Path):
        """Repositories sheet must have one data row per repo plus header."""
        inventory = await self._run_standard_discovery()
        xlsx_path = tmp_path / "report.xlsx"
        ExcelExportService.generate(inventory, xlsx_path)

        wb = load_workbook(str(xlsx_path), read_only=True)
        ws = wb["Repositories"]
        rows = list(ws.iter_rows(values_only=True))
        # row 0 = header, rows 1..n = data
        assert len(rows) == 3  # header + 2 repos
        wb.close()

    async def test_html_report_nested_subdir_created(self, tmp_path: Path):
        """ReportService must create parent directories that do not exist yet."""
        inventory = await self._run_standard_discovery()
        nested = tmp_path / "output" / "reports" / "scan.html"
        ReportService().generate(inventory, nested)
        assert nested.exists()

    async def test_excel_scan_options_in_summary(self, tmp_path: Path):
        """Summary sheet in Excel must include org name and scan profile."""
        inventory = await self._run_standard_discovery()
        xlsx_path = tmp_path / "report.xlsx"
        ExcelExportService.generate(inventory, xlsx_path)

        wb = load_workbook(str(xlsx_path), read_only=True)
        ws = wb["Summary"]
        # Summary is now 3-column: Section | Key | Value
        cell_values = {
            row[1]: row[2]
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[1] is not None and len(row) >= 3
        }
        assert cell_values.get("Organization") == _ORG
        assert cell_values.get("Scan Profile") == "standard"
        wb.close()


# ---------------------------------------------------------------------------
# Repo-limit and filter integration tests
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.asyncio
class TestRepoLimitAndFilters:
    """repo_limit and include_archived filters applied in discovery."""

    async def test_repo_limit_applied(self):
        """repo_limit=1 should keep only the first repository."""
        repos = [_make_graphql_repo("alpha"), _make_graphql_repo("beta")]
        config = _make_config(repo_limit=1)
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(repos),
            config=config,
        )
        inventory = await svc.discover()

        assert len(inventory.repositories) == 1
        assert inventory.repositories[0].name == "alpha"

    async def test_archived_excluded_when_flag_false(self):
        """include_archived=False should drop archived repos from the scan."""
        repos = [
            _make_graphql_repo("active"),
            {**_make_graphql_repo("archived"), "isArchived": True},
        ]
        config = _make_config(include_archived=False)
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(repos),
            config=config,
        )
        inventory = await svc.discover()

        assert len(inventory.repositories) == 1
        assert inventory.repositories[0].name == "active"

    async def test_archived_included_when_flag_true(self):
        """include_archived=True (default) should retain archived repos."""
        repos = [
            _make_graphql_repo("active"),
            {**_make_graphql_repo("archived"), "isArchived": True},
        ]
        config = _make_config(include_archived=True)
        svc = DiscoveryService(
            rest_client=_make_standard_rest(),
            graphql_client=_make_gql(repos),
            config=config,
        )
        inventory = await svc.discover()

        assert len(inventory.repositories) == 2
