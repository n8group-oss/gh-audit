"""Unit tests for adoption report sections (HTML + Excel)."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from gh_audit.models.adoption import (
    ActionsRunSummary,
    AdoptionInventory,
    CommitActivityInfo,
    CommunityProfileInfo,
    CopilotInfo,
    OrgCommunityHealth,
    TrafficInfo,
)
from gh_audit.models.inventory import Inventory, InventoryMetadata, InventorySummary
from gh_audit.models.repository import RepositoryInventoryItem
from gh_audit.models.user import OrgMemberSummary
from gh_audit.services.excel_export import ExcelExportService
from gh_audit.services.reporting import ReportService


def _metadata() -> InventoryMetadata:
    return InventoryMetadata(
        schema_version="2.0",
        generated_at=datetime(2026, 3, 29, tzinfo=timezone.utc),
        tool_version="0.1.0",
        organization="testorg",
        auth_method="pat",
        scan_profile="standard",
        active_categories=["adoption"],
    )


def _make_inventory(*, with_adoption: bool = False) -> Inventory:
    repo = RepositoryInventoryItem(name="repo-a", full_name="testorg/repo-a", visibility="public")
    if with_adoption:
        repo.traffic = TrafficInfo(
            views_14d=200, unique_visitors_14d=80, clones_14d=30, unique_cloners_14d=15
        )
        repo.community_profile = CommunityProfileInfo(
            health_percentage=71, has_readme=True, has_license=True
        )
        repo.commit_activity_90d = CommitActivityInfo(total_commits=50, active_weeks=8)
        repo.actions_run_summary = ActionsRunSummary(
            total_runs_90d=100, by_conclusion={"success": 85, "failure": 10, "cancelled": 5}
        )

    adoption = None
    if with_adoption:
        adoption = AdoptionInventory(
            copilot=CopilotInfo(
                total_seats=50, active_seats=40, suggestions_count=1200, acceptances_count=800
            ),
            org_community_health=OrgCommunityHealth(
                repos_with_readme=1, repos_with_license=1, average_health_percentage=71.0
            ),
        )

    return Inventory(
        metadata=_metadata(),
        summary=InventorySummary(total_repos=1),
        repositories=[repo],
        users=OrgMemberSummary(total=1, admins=0, members=1),
        adoption=adoption,
    )


class TestHTMLAdoption:
    def test_adoption_section_present(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_adoption=True)
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        assert "Adoption" in content
        assert "Copilot" in content
        assert "Community Health" in content

    def test_adoption_section_absent(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_adoption=False)
        output = tmp_path / "report.html"
        ReportService().generate(inv, output)
        content = output.read_text()
        # "Adoption" should not appear as a section heading
        assert "Adoption</h2>" not in content


class TestExcelAdoption:
    def test_copilot_sheet(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_adoption=True)
        output = tmp_path / "report.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Copilot" in wb.sheetnames
        ws = wb["Copilot"]
        assert ws["A1"].value == "Metric"
        assert ws["A2"].value == "Total Seats"
        assert ws["B2"].value == 50

    def test_traffic_sheet(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_adoption=True)
        output = tmp_path / "report.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Traffic" in wb.sheetnames
        ws = wb["Traffic"]
        assert ws["A1"].value == "Repository"
        assert ws["A2"].value == "repo-a"
        assert ws["B2"].value == 200

    def test_community_health_sheet(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_adoption=True)
        output = tmp_path / "report.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Community Health" in wb.sheetnames
        ws = wb["Community Health"]
        assert ws["A2"].value == "repo-a"
        assert ws["B2"].value == 71

    def test_actions_runs_sheet(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_adoption=True)
        output = tmp_path / "report.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Actions Runs" in wb.sheetnames
        ws = wb["Actions Runs"]
        assert ws["A2"].value == "repo-a"
        assert ws["B2"].value == 100

    def test_no_adoption_sheets_without_data(self, tmp_path: Path) -> None:
        inv = _make_inventory(with_adoption=False)
        output = tmp_path / "report.xlsx"
        ExcelExportService.generate(inv, output)
        wb = load_workbook(output)
        assert "Copilot" not in wb.sheetnames
        assert "Traffic" not in wb.sheetnames
        assert "Community Health" not in wb.sheetnames
        assert "Actions Runs" not in wb.sheetnames
